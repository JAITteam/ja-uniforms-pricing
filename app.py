# ===== LOAD ENVIRONMENT VARIABLES (ONCE!) =====
from dotenv import load_dotenv
load_dotenv()

# ===== STANDARD LIBRARY =====
import os
import re
import csv
import json
import random
import string
import html
import secrets
import logging
from io import StringIO
from datetime import datetime, timedelta
from logging.handlers import RotatingFileHandler

# ===== THIRD PARTY =====
import pytz
import pandas as pd
from flask import Flask, render_template, request, jsonify, redirect, url_for, send_file, Response, flash, session
from flask_mail import Mail, Message
from flask_migrate import Migrate
from flask_login import login_user, logout_user, current_user, login_required
from flask_wtf.csrf import CSRFProtect
from sqlalchemy.exc import IntegrityError
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from sqlalchemy import func
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash
from sqlalchemy.orm import joinedload

# ===== LOCAL IMPORTS =====
from config import Config
from database import db
from auth import init_auth, admin_required, login_required_custom, role_required, get_user_permissions
from models import (
    Style, Fabric, User, FabricVendor, Notion, NotionVendor,
    LaborOperation, CleaningCost, StyleFabric, StyleNotion,
    StyleLabor, Color, StyleColor, Variable, StyleVariable,
    SizeRange, GlobalSetting, StyleImage, VerificationCode, AuditLog
)



# ===== HELPER FUNCTIONS =====
def validate_password(password):
    """Validate password strength"""
    if not password:
        return False, "Password is required"
    if len(password) < 8:
        return False, "Password must be at least 8 characters"
    if not re.search(r'[A-Z]', password):
        return False, "Password must contain at least one uppercase letter"
    if not re.search(r'[a-z]', password):
        return False, "Password must contain at least one lowercase letter"
    if not re.search(r'\d', password):
        return False, "Password must contain at least one number"
    return True, "Valid"

def validate_image_content(file_header):
    """
    Validate image file by checking magic bytes (file signature).
    Returns detected type ('jpeg', 'png', 'gif') or None if invalid.
    """
    # Magic bytes for common image formats
    # JPEG: starts with FF D8 FF
    # PNG: starts with 89 50 4E 47 0D 0A 1A 0A
    # GIF: starts with GIF87a or GIF89a
    
    if file_header[:3] == b'\xff\xd8\xff':
        return 'jpeg'
    elif file_header[:8] == b'\x89PNG\r\n\x1a\n':
        return 'png'
    elif file_header[:6] in (b'GIF87a', b'GIF89a'):
        return 'gif'
    else:
        return None

# Initialize Mail
mail = Mail()


def setup_logging(app):
    """Configure application logging"""
    # Create logs folder if it doesn't exist
    if not os.path.exists('logs'):
        os.makedirs('logs')
    
    # Main log file - captures everything (with UTF-8 encoding for emojis)
    file_handler = RotatingFileHandler(
        'logs/ja_uniforms.log',
        maxBytes=10 * 1024 * 1024,  # 10MB max size
        backupCount=10,              # Keep 10 backup files
        encoding='utf-8'             # Support emojis and special characters
    )
    file_handler.setFormatter(logging.Formatter(
        '[%(asctime)s] %(levelname)s in %(module)s: %(message)s'
    ))
    file_handler.setLevel(logging.INFO)
    
    # Error-only log file - easier to find critical issues
    error_handler = RotatingFileHandler(
        'logs/ja_uniforms_errors.log',
        maxBytes=10 * 1024 * 1024,
        backupCount=10,
        encoding='utf-8'             # Support emojis and special characters
    )
    error_handler.setFormatter(logging.Formatter(
        '[%(asctime)s] %(levelname)s in %(pathname)s:%(lineno)d\n%(message)s\n'
    ))
    error_handler.setLevel(logging.ERROR)
    
    # Console handler with UTF-8 support (fixes Windows emoji issue)
    console_handler = logging.StreamHandler()
    console_handler.setFormatter(logging.Formatter(
        '[%(asctime)s] %(levelname)s in %(module)s: %(message)s'
    ))
    console_handler.setLevel(logging.INFO)
    
    # Add handlers to app
    app.logger.addHandler(file_handler)
    app.logger.addHandler(error_handler)
    app.logger.addHandler(console_handler)
    app.logger.setLevel(logging.INFO)
    
    app.logger.info('=== J.A. Uniforms Application Started ===')


# Initialize Flask app
app = Flask(__name__)
app.config['REMEMBER_COOKIE_DURATION'] = timedelta(days=30)  # 30 days instead of 365
setup_logging(app)
# Rate Limiter - prevents abuse
limiter = Limiter(
    app=app,
    key_func=get_remote_address,
    default_limits=["2000 per day", "500 per hour"],
    storage_uri="memory://"
)
app.config.from_object(Config)

app.config['MAIL_SERVER'] = os.getenv('MAIL_SERVER', 'smtp.gmail.com')
app.config['MAIL_PORT'] = int(os.getenv('MAIL_PORT', 587))
app.config['MAIL_USE_TLS'] = True
app.config['MAIL_USERNAME'] = os.getenv('MAIL_USERNAME')
app.config['MAIL_PASSWORD'] = os.getenv('MAIL_PASSWORD')  # ✅ SECURE
app.config['MAIL_DEFAULT_SENDER'] = os.getenv('MAIL_DEFAULT_SENDER')
# =========================================
mail.init_app(app)

@app.template_filter('fromjson')
def fromjson_filter(value):
    try:
        return json.loads(value) if value else {}
    except:
        return {}


# Initialize CSRF protection
csrf = CSRFProtect(app)

init_auth(app)

@app.context_processor
def inject_permissions():
    """Make user permissions available to all templates"""
    if current_user.is_authenticated:
        is_admin = current_user.role == 'admin'
        return {
            'permissions': get_user_permissions(),
            'user_role': current_user.role,
            'can_edit_styles': is_admin,
            'can_delete_styles': is_admin,
            'can_create_styles': is_admin,
            'can_duplicate_styles': is_admin,
            'can_favorite_styles': is_admin,
            'is_admin': is_admin
        }
    return {
        'permissions': get_user_permissions(),
        'user_role': None,
        'can_edit_styles': False,
        'can_delete_styles': False,
        'can_create_styles': False,
        'can_duplicate_styles': False,
        'can_favorite_styles': False,
        'is_admin': False
    }


UPLOAD_FOLDER = 'static/img'
ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}
ALLOWED_EXCEL_EXTENSIONS = {'xlsx', 'xls'}
MAX_FILE_SIZE = 16 * 1024 * 1024  # 16MB

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def allowed_file(filename, allowed_extensions):
    """Check if file extension is allowed"""
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in allowed_extensions

# Initialize database with app
db.init_app(app)
migrate = Migrate(app, db)

# ===== HELPER FUNCTIONS (already provided in your document) =====

def generate_verification_code():
    """Generate a 6-digit verification code"""
    return ''.join(random.choices(string.digits, k=6))

def require_admin_for_write():
    """Check if current request is a write operation and require admin"""
    if request.method in ['POST', 'PUT', 'DELETE', 'PATCH']:
        if not current_user.is_authenticated:
            return jsonify({'error': 'Authentication required'}), 401
        if not current_user.is_admin():
            return jsonify({'error': 'Admin access required'}), 403
    return None


def send_verification_email(email, code):
    """Send verification code to email"""
    msg = Message(
        'J.A. Uniforms - Email Verification Code',
        recipients=[email]
    )
    msg.body = f'''
Hello,

Your verification code for J.A. Uniforms Pricing Tool is:

{code}

This code will expire in 10 minutes.

If you didn't request this code, please ignore this email.

Best regards,
J.A. Uniforms Team
    '''
    
    msg.html = f'''
<html>
<body style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
    <div style="max-width: 600px; margin: 0 auto; padding: 20px;">
        <div style="background: linear-gradient(135deg, #3498db 0%, #2980b9 100%); padding: 30px; border-radius: 10px; text-align: center;">
            <h1 style="color: white; margin: 0;">J.A. Uniforms</h1>
            <p style="color: white; margin: 10px 0 0 0;">Pricing Management System</p>
        </div>
        
        <div style="padding: 30px; background: #f9f9f9; border-radius: 10px; margin-top: 20px;">
            <h2 style="color: #2c3e50;">Email Verification Code</h2>
            <p>Hello,</p>
            <p>Your verification code for J.A. Uniforms Pricing Tool is:</p>
            
            <div style="background: white; padding: 20px; border-radius: 8px; text-align: center; margin: 20px 0;">
                <h1 style="color: #3498db; font-size: 36px; letter-spacing: 8px; margin: 0;">{code}</h1>
            </div>
            
            <p style="color: #666; font-size: 14px;">This code will expire in 10 minutes.</p>
            <p style="color: #666; font-size: 14px;">If you didn't request this code, please ignore this email.</p>
        </div>
        
        <div style="text-align: center; margin-top: 20px; color: #999; font-size: 12px;">
            <p>© 2025 J.A. Uniforms. All rights reserved.</p>
        </div>
    </div>
</body>
</html>
    '''
    
    try:
        mail.send(msg)
        return True
    except Exception as e:
        app.logger.error(f"Error sending email: {e}")
        return False

# ===== VALIDATION HELPER FUNCTIONS =====

def validate_positive_number(value, field_name, required=True, allow_zero=True):
    """Validate that a value is a positive number"""
    if value is None or value == '':
        if required:
            return None, f"{field_name} is required"
        return 0 if allow_zero else None, None
    
    try:
        num = float(value)
        if num < 0:
            return None, f"{field_name} cannot be negative"
        if not allow_zero and num == 0:
            return None, f"{field_name} must be greater than zero"
        return num, None
    except (ValueError, TypeError):
        return None, f"{field_name} must be a valid number"

def validate_positive_integer(value, field_name, required=True):
    """Validate that a value is a positive integer"""
    if value is None or value == '':
        if required:
            return None, f"{field_name} is required"
        return 0, None
    
    try:
        num = int(float(value))
        if num < 0:
            return None, f"{field_name} cannot be negative"
        return num, None
    except (ValueError, TypeError):
        return None, f"{field_name} must be a valid whole number"
def safe_margin_calculation(cost, margin_percent):
    """
    Calculate sale price from cost and margin, preventing division by zero.
    Caps margin at 99% to avoid unrealistic prices.
    """
    if cost is None or cost <= 0:
        return 0
    margin_decimal = (margin_percent or 60) / 100.0
    if margin_decimal >= 0.99:
        margin_decimal = 0.99
    if margin_decimal < 0:
        margin_decimal = 0
    return round(cost / (1 - margin_decimal), 2)

def get_next_fabric_code():
    """Generate next sequential fabric code like T1, T2, T3..."""
    existing = Fabric.query.with_entities(Fabric.fabric_code).all()
    max_num = 0
    for (code,) in existing:
        if code and code.upper().startswith('T'):
            try:
                num = int(code[1:])
                if num > max_num:
                    max_num = num
            except ValueError:
                pass
    return f"T{max_num + 1}"
        
def validate_required_string(value, field_name, max_length=200):
    """Validate that a value is a non-empty string"""
    if value is None or str(value).strip() == '':
        return None, f"{field_name} is required"
    
    value = str(value).strip()
    if len(value) > max_length:
        return None, f"{field_name} cannot exceed {max_length} characters"
    
    return value, None

def validate_choice(value, field_name, choices):
    """Validate that a value is one of the allowed choices"""
    if value not in choices:
        return None, f"{field_name} must be one of: {', '.join(choices)}"
    return value, None


def validate_email(email):
    """Validate email format"""
    if not email:
        return None, "Email is required"
    
    email = email.strip().lower()
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    if not re.match(pattern, email):
        return None, "Invalid email format"
    
    return email, None


def validate_required_field(value, field_name):
    """Validate that a required field is not empty"""
    if not value or not str(value).strip():
        return False, f"{field_name} is required"
    return True, str(value).strip()
    
def validate_percentage(value, field_name, max_value=99):
    """Validate that a percentage is between 0 and max_value (default 99 to prevent division by zero)"""
    try:
        num = float(value)
        if num < 0 or num > max_value:
            return False, f"{field_name} must be between 0 and {max_value}"
        return True, num
    except (ValueError, TypeError):
        return False, f"{field_name} must be a valid number"

def validate_string_length(value, field_name, max_length):
    """Validate string length"""
    if len(str(value)) > max_length:
        return False, f"{field_name} too long (max {max_length} characters)"
    return True, str(value)


def sanitize_search_query(query, max_length=100):
    """
    Sanitize user input for safe use in SQL LIKE queries.
    """
    if not query:
        return None, None
    
    # Strip and check minimum length
    query = str(query).strip()
    if len(query) < 2:
        return None, None
    
    # Escape special SQL LIKE characters
    sanitized = (
        query
        .replace('\\', '\\\\')  # Escape backslash first
        .replace('%', '\\%')    # Escape percent
        .replace('_', '\\_')    # Escape underscore
    )
    
    # Limit length
    sanitized = sanitized[:max_length]
    
    # Create search pattern
    search_pattern = f'%{sanitized}%'
    
    return sanitized, search_pattern

# ===== END OF VALIDATION HELPERS =====
@app.errorhandler(429)
def ratelimit_handler(e):
    """Handle rate limit exceeded"""
    if request.path.startswith('/api/'):
        return jsonify({
            "success": False,
            "error": "Too many requests. Please slow down."
        }), 429   
    flash("Too many requests. Please wait a moment and try again.", "warning")
    return redirect(request.referrer or url_for('index'))

# ===== YOUR ROUTES START HERE =====

def log_audit(action, item_type, item_id=None, item_name=None, 
              old_values=None, new_values=None, affected_styles_count=0, details=None):
    """Log an audit entry for tracking changes"""
    try:
        user_name = 'System'
        user_email = None
        user_id = None
        
        if current_user.is_authenticated:
            user_name = current_user.email
            user_email = current_user.email
            user_id = current_user.id
        
        log_entry = AuditLog(
            user_id=user_id,
            user_name=user_name,
            user_email=user_email,
            action=action,
            item_type=item_type,
            item_id=item_id,
            item_name=item_name,
            old_values=json.dumps(old_values) if old_values else None,
            new_values=json.dumps(new_values) if new_values else None,
            affected_styles_count=affected_styles_count,
            details=details
        )
        db.session.add(log_entry)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Audit log error: {e}")

@app.route('/audit-logs')
@login_required
@admin_required
def audit_logs():
    page = request.args.get('page', 1, type=int)
    per_page = 50
    
    # Get filter parameters
    item_type = request.args.get('item_type', '')
    action = request.args.get('action', '')
    search = request.args.get('search', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    query = AuditLog.query
    
    # Apply filters
    if item_type:
        query = query.filter(AuditLog.item_type == item_type)
    if action:
        query = query.filter(AuditLog.action == action)
    if search:
        sanitized, search_pattern = sanitize_search_query(search)
        if sanitized:
            query = query.filter(AuditLog.item_name.ilike(search_pattern, escape='\\'))
    if date_from:
        try:
            date_from_obj = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(AuditLog.timestamp >= date_from_obj)
        except ValueError:
            pass
    if date_to:
        try:
            date_to_obj = datetime.strptime(date_to, '%Y-%m-%d') + timedelta(days=1)
            query = query.filter(AuditLog.timestamp < date_to_obj)
        except ValueError:
            pass
    
    logs = query.order_by(AuditLog.timestamp.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return render_template('audit_logs.html', 
                          logs=logs, 
                          current_item_type=item_type, 
                          current_action=action,
                          current_search=search,
                          current_date_from=date_from,
                          current_date_to=date_to)

@app.route('/api/send-verification-code', methods=['POST'])
@limiter.limit("3 per minute") 
@limiter.limit("10 per hour") 
def send_verification_code():
    """API endpoint to send verification code"""
    data = request.get_json()
    email = data.get('email', '').strip().lower()
    
    if not email:
        return jsonify({'success': False, 'error': 'Email is required'}), 400
    
    # Validate email domain
    if not email.endswith('@jauniforms.com'):
        return jsonify({'success': False, 'error': 'Only company emails are allowed'}), 400
    
    # Check if user exists
    if User.query.filter_by(username=email).first():
        return jsonify({'success': False, 'error': 'Email already registered'}), 400
    
    # Generate and store verification code
    code = generate_verification_code()
    try:
        # Delete any existing code for this email and create new one atomically
        VerificationCode.query.filter_by(email=email).delete()
        # Save to database
        verification = VerificationCode(
            email=email,
            code=code,
            password_hash='pending',  # Will be set in register route
            first_name='',
            last_name='',
            expires_at=datetime.now() + timedelta(minutes=10)
        )
        db.session.add(verification)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        # Race condition occurred - another request just created a code
        app.logger.warning(f"Race condition in verification code for {email}")
        
        # Check if a valid code already exists and resend it
        existing_code = VerificationCode.query.filter_by(email=email).first()
        if existing_code and not existing_code.is_expired():
            # Resend the existing code
            if send_verification_email(email, existing_code.code):
                return jsonify({'success': True, 'message': 'Verification code sent successfully'}), 200
            else:
                return jsonify({'success': False, 'error': 'Failed to send verification email. Please try again.'}), 500
        else:
            # No valid code exists, ask user to try again
            return jsonify({'success': False, 'error': 'Please try again in a moment.'}), 429
    
    # Send email
    if send_verification_email(email, code):
        return jsonify({'success': True, 'message': 'Verification code sent successfully'}), 200
    else:
        return jsonify({'success': False, 'error': 'Failed to send email'}), 500
    
@app.route('/api/verify-code', methods=['POST'])
@limiter.limit("5 per minute")
def verify_code():
    """Verify the email verification code"""
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        code = data.get('code', '').strip()
        
        # Get from database
        verification = VerificationCode.query.filter_by(email=email).first()
        
        if not verification:
            return jsonify({'success': False, 'error': 'No verification code found'}), 200
        
        # Check if expired
        if verification.is_expired():
            db.session.delete(verification)
            db.session.commit()
            return jsonify({'success': False, 'error': 'Code expired. Please register again'}), 200
        
        # Check if code matches
        if verification.code != code:
            return jsonify({'success': False, 'error': 'Invalid code'}), 200
        
        # Create user
        user = User(
            username=email,
            email=email,
            first_name=verification.first_name,
            last_name=verification.last_name,
            full_name=f"{verification.first_name or ''} {verification.last_name or ''}".strip()
        )
        user.password_hash = verification.password_hash
        # Assign role based on email

        admin_emails = os.getenv('ADMIN_EMAILS', 'it@jauniforms.com').split(',')
        if email in [e.strip() for e in admin_emails]:
            user.role = 'admin'
        else:
            user.role = 'user'
        
        db.session.add(user)
        
        # Delete verification record
        db.session.delete(verification)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Verification successful',
            'redirect': url_for('login')
        })
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Verification error: {e}")
        return jsonify({'success': False, 'error': 'Verification failed'}), 200

@app.route('/api/resend-verification-code', methods=['POST'])
@limiter.limit("2 per minute")
@limiter.limit("5 per hour")
def resend_verification_code():
    """Resend verification code"""
    try:
        data = request.get_json()
        email = data.get('email', '').strip().lower()
        
        # Get from database
        verification = VerificationCode.query.filter_by(email=email).first()
        
        if not verification:
            return jsonify({'success': False, 'error': 'No registration found'}), 200
        
        # Generate new code
        new_code = generate_verification_code()
        verification.code = new_code
        verification.expires_at = datetime.now() + timedelta(minutes=10)
        db.session.commit()
        
        # Send new email
        if send_verification_email(email, new_code):
            return jsonify({'success': True, 'message': 'New code sent'}), 200
        else:
            return jsonify({'success': False, 'error': 'Error sending email'}), 200
            
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Resend verification error: {e}")
        return jsonify({'success': False, 'error': 'Failed to resend code'}), 200

# ===== AUTHENTICATION ROUTES =====
@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("10 per minute") 
def login():
    # Don't redirect if already logged in - let them see the login page
    # if current_user.is_authenticated:
    #     return redirect(url_for('index'))
    
    if request.method == 'POST':
        email = request.form.get('username', '').strip().lower()
        password = request.form.get('password')
        
        if not email or not email.endswith('@jauniforms.com'):
            flash('Only company emails (@jauniforms.com) are allowed', 'danger')
            return render_template('login.html')
        
        user = User.query.filter_by(username=email).first()
        
        if user and user.check_password(password):
            login_user(user, remember=request.form.get('remember'))
            
            # Don't show welcome message if user needs to change password
            if not user.must_change_password:
                flash(f'Welcome back, {user.username}!', 'success')
            
            next_page = request.args.get('next')
            return redirect(next_page or url_for('index'))
        else:
            flash('Invalid email or password', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
@login_required
def logout():
    """Logout route"""
    logout_user()
    # Clear all existing flash messages
    session.pop('_flashes', None)
    # Add only logout message
    flash('You have been logged out successfully.', 'info')
    return redirect(url_for('login'))

@app.route('/register', methods=['GET', 'POST'])
@limiter.limit("10 per minute") 
def register():
    """Main registration page - single page with modal verification"""
    if current_user.is_authenticated:
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        try:
            # Get form data
            first_name = request.form.get('firstName', '').strip()
            last_name = request.form.get('lastName', '').strip()
            email = request.form.get('email', '').strip().lower()
            password = request.form.get('password')

            # Name length validation
            if not first_name or len(first_name) > 50:
                return jsonify({'success': False, 'error': 'First name is required and must be 50 characters or less'}), 200
            if not last_name or len(last_name) > 50:
                return jsonify({'success': False, 'error': 'Last name is required and must be 50 characters or less'}), 200
            
            # Validation
            if not email or not email.endswith('@jauniforms.com'):
                return jsonify({'success': False, 'error': 'Only company emails (@jauniforms.com) are allowed'}), 200
            
            is_valid, message = validate_password(password)
            if not is_valid:
                return jsonify({'success': False, 'error': message}), 200
            
            if User.query.filter_by(username=email).first():
                return jsonify({'success': False, 'error': 'Email already registered'}), 200
            
            # Generate code
            code = generate_verification_code()
            
            # Delete any existing code for this email
            VerificationCode.query.filter_by(email=email).delete()
            
            # Store in database (survives restart!)
            verification = VerificationCode(
                email=email,
                code=code,
                password_hash=generate_password_hash(password),
                first_name=first_name,
                last_name=last_name,
                expires_at=datetime.now() + timedelta(minutes=10)
            )
            db.session.add(verification)
            db.session.commit()
            
            # Send verification email
            if send_verification_email(email, code):
                app.logger.info(f"Verification code sent to {email}")
                return jsonify({'success': True, 'message': 'Verification code sent'}), 200
            else:
                return jsonify({'success': False, 'error': 'Error sending email'}), 200
                
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Registration error: {e}")
            return jsonify({'success': False, 'error': 'Registration failed'}), 200
    
    return render_template('register.html')
# ===== USER MANAGEMENT ROUTES (ADMIN ONLY) =====

@app.route('/admin/users')
@admin_required
def manage_users():
    """Admin-only user management page"""
    users = User.query.order_by(User.created_at.desc()).all()
    return render_template('admin_users.html', users=users)

@app.route('/api/users/<int:user_id>', methods=['GET'])
@admin_required
def get_user(user_id):
    """Get user details"""
    user = User.query.get_or_404(user_id)
    return jsonify({
        'id': user.id,
        'email': user.email,
        'first_name': user.first_name or '',
        'last_name': user.last_name or '',
        'role': user.role,
        'is_active': user.is_active,
        'created_at': user.created_at.isoformat() if user.created_at else None
    })

@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
@admin_required
def update_user(user_id):
    """Update user details"""
    try:
        user = User.query.get_or_404(user_id)
        data = request.json
        
        # CRITICAL: Prevent removing the last admin
        if 'role' in data and data['role'] == 'user' and user.role == 'admin':
            admin_count = User.query.filter_by(role='admin').count()
            
            if admin_count <= 1:
                app.logger.warning("Attempted to remove last admin - blocked")
                return jsonify({
                    'success': False, 
                    'error': 'Cannot change the last admin to user. The system must have at least one admin.'
                }), 400
        
        # Update email if provided and check for duplicates
        if 'email' in data and data['email'] != user.email:
            existing = User.query.filter_by(email=data['email']).first()
            if existing and existing.id != user_id:
                return jsonify({'success': False, 'error': 'Email already exists'}), 400
            user.email = data['email']
            user.username = data['email']
        
        if 'first_name' in data:
            user.first_name = data['first_name']
        if 'last_name' in data:
            user.last_name = data['last_name']
        if 'role' in data and data['role'] in ['admin', 'user']:
            user.role = data['role']
        if 'is_active' in data:
            user.is_active = data['is_active']
        
        if user.first_name and user.last_name:
            user.full_name = f"{user.first_name} {user.last_name}"
        elif user.first_name:
            user.full_name = user.first_name
        
        db.session.commit()
        app.logger.info(f"User {user_id} updated successfully")
        
        return jsonify({'success': True, 'message': 'User updated successfully'})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error updating user {user_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    
@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@limiter.limit("5 per minute")
@login_required
@admin_required
def delete_user(user_id):
    """Delete user"""
    try:
        user = User.query.get_or_404(user_id)
        
        # Prevent deleting yourself
        if user.id == current_user.id:
            return jsonify({'success': False, 'error': 'Cannot delete your own account'}), 400
        
        # CRITICAL: Prevent deleting the last admin
        if user.role == 'admin':
            admin_count = User.query.filter_by(role='admin').count()
            
            if admin_count <= 1:
                app.logger.warning("Attempted to delete last admin - blocked")
                return jsonify({
                    'success': False, 
                    'error': 'Cannot delete the last admin. The system must have at least one admin.'
                }), 400
        
        db.session.delete(user)
        db.session.commit()
        app.logger.info(f"User {user_id} deleted successfully")
        
        return jsonify({'success': True, 'message': 'User deleted successfully'})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error deleting user {user_id}: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500
    

# ===== PASSWORD RESET ROUTES =====

@app.route('/api/users/<int:user_id>/reset-password', methods=['POST'])
@limiter.limit("5 per minute")
@login_required
@admin_required
def admin_reset_password(user_id):
    """
    Admin resets a user's password to a temporary one.
    User will be forced to change it on next login.
    """
    try:
        user = User.query.get_or_404(user_id)
        
        # Don't allow resetting your own password this way
        if user.id == current_user.id:
            return jsonify({
                'success': False, 
                'error': 'Cannot reset your own password. Use the change password feature instead.'
            }), 400
        
        # Generate a secure temporary password (12 characters)
        temp_password = secrets.token_urlsafe(9)[:12]
        
        # Set the temporary password
        user.set_password(temp_password)
        user.must_change_password = True
        user.temp_password_created_at = datetime.now()
        
        db.session.commit()
        
        app.logger.info(f"Admin {current_user.email} reset password for user {user.email}")
        
        return jsonify({
            'success': True,
            'temp_password': temp_password,
            'message': f'Temporary password generated for {user.get_full_name()}',
            'user_email': user.email
        }), 200
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Password reset error: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    """
    Force password change page.
    Shown when user.must_change_password is True.
    """
    if request.method == 'GET':
        # If user doesn't need to change password, redirect to dashboard
        if not current_user.must_change_password:
            return redirect(url_for('index'))
        return render_template('change_password.html')
    
    # POST - Handle password change
    try:
        current_password = request.form.get('current_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        
        # Validate current password
        if not current_user.check_password(current_password):
            flash('Current password is incorrect.', 'danger')
            return render_template('change_password.html')
        
        # Validate new password
        is_valid, message = validate_password(new_password)
        if not is_valid:
            flash(message, 'danger')
            return render_template('change_password.html')
        
        # Check passwords match
        if new_password != confirm_password:
            flash('New passwords do not match.', 'danger')
            return render_template('change_password.html')
        
        # Don't allow same password as current
        if current_user.check_password(new_password):
            flash('New password must be different from your current password.', 'danger')
            return render_template('change_password.html')
        
        # Update password
        current_user.set_password(new_password)
        current_user.must_change_password = False
        current_user.temp_password_created_at = None
        
        db.session.commit()
        
        app.logger.info(f"User {current_user.email} changed password successfully")
        
        flash('Password changed successfully! Welcome to J.A. Uniforms.', 'success')
        return redirect(url_for('index'))
        
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Change password error: {e}")
        flash('Error changing password. Please try again.', 'danger')
        return render_template('change_password.html')

@app.template_filter('time_ago')
def time_ago_filter(dt):
    """Calculate time ago using system time"""
    if dt is None:
        return 'N/A'
    
    now = datetime.now()
    diff = now - dt
    
    seconds = diff.total_seconds()
    
    if seconds < 0:
        return 'Just now'
    elif seconds < 60:
        return 'Just now'
    elif seconds < 3600:
        minutes = int(seconds / 60)
        return f'{minutes} min ago'
    elif seconds < 86400:
        hours = int(seconds / 3600)
        return f'{hours} hour{"s" if hours != 1 else ""} ago'
    elif seconds < 604800:
        days = int(seconds / 86400)
        return f'{days} day{"s" if days != 1 else ""} ago'
    else:
        return dt.strftime('%b %d, %Y')

# ===== PASSWORD CHANGE INTERCEPTOR =====
@app.before_request
def check_password_change_required():
    """
    Intercept requests and redirect to change password if required.
    """
    # Skip for static files, login, logout, and change-password routes
    allowed_endpoints = ['login', 'logout', 'change_password', 'static', None]
    
    if request.endpoint in allowed_endpoints:
        return None
    
    # Check if user is logged in and must change password
    if current_user.is_authenticated and getattr(current_user, 'must_change_password', False):
        if request.endpoint != 'change_password':
            flash('Please change your temporary password to continue.', 'warning')
            return redirect(url_for('change_password'))
    

# ===== MAIN APPLICATION ROUTES =====

@app.route('/')
@login_required
def index():
    """Dashboard with real-time stats"""
    
    # Get real counts
    total_styles = Style.query.count()
    total_fabrics = Fabric.query.count()
    total_notions = Notion.query.count()
    total_fabric_vendors = FabricVendor.query.count()
    total_notion_vendors = NotionVendor.query.count()
    
    # New styles this week
    
    one_week_ago = datetime.now() - timedelta(days=7)
    new_this_week = Style.query.filter(Style.created_at >= one_week_ago).count()
    
    recent_styles = Style.query.order_by(Style.updated_at.desc()).limit(4).all()

    # Calculate analytics with eager loading (faster!)
    styles = Style.query.options(
        joinedload(Style.style_fabrics).joinedload(StyleFabric.fabric),
        joinedload(Style.style_notions).joinedload(StyleNotion.notion),
        joinedload(Style.style_labor).joinedload(StyleLabor.labor_operation)
    ).all()

    # Cache label cost (1 query instead of N)
    label_setting = GlobalSetting.query.filter_by(setting_key='avg_label_cost').first()
    label_cost = label_setting.setting_value if label_setting else 0.20

    # Average cost per style
    if styles:
        costs = [s.get_total_fabric_cost() + s.get_total_notion_cost() + s.get_total_labor_cost() + label_cost for s in styles]
        total_cost = sum(costs)
        avg_cost = total_cost / len(styles)
        min_cost = min(costs) if costs else 0
        max_cost = max(costs) if costs else 0
        price_range = f"${min_cost:.0f}-${max_cost:.0f}"
    else:
        avg_cost = 0
        price_range = "$0-$0"

    # Top fabric (most used)
    from sqlalchemy import func
    top_fabric_query = db.session.query(
        Fabric.name, 
        func.count(StyleFabric.fabric_id).label('count')
    ).join(
        StyleFabric, Fabric.id == StyleFabric.fabric_id
    ).group_by(
        Fabric.name
    ).order_by(
        func.count(StyleFabric.fabric_id).desc()
    ).first()
    
    top_fabric = top_fabric_query[0] if top_fabric_query else "N/A"
    
    return render_template('dashboard.html',
                         total_styles=total_styles,
                         total_fabrics=total_fabrics,
                         total_notions=total_notions,
                         total_fabric_vendors=total_fabric_vendors,
                         total_notion_vendors=total_notion_vendors,
                         recent_styles=recent_styles,
                         avg_cost=avg_cost,
                         price_range=price_range,
                         top_fabric=top_fabric,
                         new_this_week=new_this_week)


@app.route('/style/<vendor_style>')
@login_required
def view_style(vendor_style):
    style = Style.query.filter_by(vendor_style=vendor_style).first()
    if style:
        return redirect(url_for('style_view') + f'?vendor_style={vendor_style}')
    return f"<h1>Style '{vendor_style}' not found</h1><a href='/'>Back to Search</a>"

# ===== ADMIN ROUTES (Future expansion) =====
@app.route('/api/recent-styles')
@login_required
def api_recent_styles():
    styles = Style.query.order_by(Style.updated_at.desc()).limit(5).all()
    return jsonify([{
        'id': s.id,
        'vendor_style': s.vendor_style,
        'style_name': s.style_name,
        'updated_at': s.updated_at.isoformat() if s.updated_at else None
    } for s in styles])

@app.route('/api/all-styles-for-export')
@login_required
def api_all_styles_for_export():
    """Get all styles for export modal"""
    styles = Style.query.order_by(Style.vendor_style).all()
    return jsonify([{
        'id': s.id,
        'vendor_style': s.vendor_style,
        'style_name': s.style_name,
        'gender': s.gender or 'N/A',
        'cost': s.get_total_cost()
    } for s in styles])

@app.route('/api/dashboard-stats')
@login_required
def api_dashboard_stats():
    total_styles = Style.query.count()
    
    # Eager load to avoid N+1 queries
    styles = Style.query.options(
        joinedload(Style.style_fabrics).joinedload(StyleFabric.fabric),
        joinedload(Style.style_notions).joinedload(StyleNotion.notion),
        joinedload(Style.style_labor).joinedload(StyleLabor.labor_operation)
    ).all()
    
    # Cache label cost
    label_setting = GlobalSetting.query.filter_by(setting_key='avg_label_cost').first()
    label_cost = label_setting.setting_value if label_setting else 0.20
    
    if styles:
        total_cost = sum(s.get_total_fabric_cost() + s.get_total_notion_cost() + s.get_total_labor_cost() + label_cost for s in styles)
        avg_cost = total_cost / len(styles)
    else:
        avg_cost = 0

    return jsonify({
        'total_styles': total_styles,
         'avg_cost': round(avg_cost, 2)
    })

@app.route('/api/style/<int:style_id>/upload-image', methods=['POST'])
@admin_required 
def upload_style_image(style_id):
    """Upload an image for a style - with size and MIME validation"""
    
    # Check if style exists
    style = Style.query.get_or_404(style_id)
    
    # Check if file is in request
    if 'image' not in request.files:
        return jsonify({'error': 'No file provided'}), 400
    
    file = request.files['image']
    
    # Check if filename is empty
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    # ✅ NEW: Check file size BEFORE processing
    file.seek(0, 2)  # Seek to end of file
    file_size = file.tell()  # Get current position (= file size)
    file.seek(0)  # Reset to beginning
    
    if file_size > MAX_FILE_SIZE:
        max_mb = MAX_FILE_SIZE // (1024 * 1024)
        return jsonify({'error': f'File too large. Maximum size is {max_mb}MB'}), 400
    
    if file_size == 0:
        return jsonify({'error': 'File is empty'}), 400
    
    # Check extension first (quick check)
    if not allowed_file(file.filename, ALLOWED_IMAGE_EXTENSIONS):
        return jsonify({'error': 'Invalid file type. Allowed: png, jpg, jpeg, gif'}), 400
    
    
    # ✅ Validate actual file content (magic byte check)
    # Read first 16 bytes to determine real file type
    header = file.read(16)
    file.seek(0)  # Reset to beginning
    
    # Check magic bytes to verify it's a real image
    detected_type = validate_image_content(header)
    
    if detected_type is None:
        app.logger.warning(f"Upload rejected: File extension was valid but content is not a valid image")
        return jsonify({'error': 'Invalid image file. File content does not match a valid image format.'}), 400
    
    # ✅ All validations passed - now save the file
    
    # Create unique filename with timestamp
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    original_filename = secure_filename(file.filename)
    filename = f"style_{style_id}_{timestamp}_{original_filename}"
    
    # Ensure upload directory exists
    os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
    
    # Save file to disk
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    file.save(filepath)
    
    # Check if this should be the primary image (first image for this style)
    is_primary = StyleImage.query.filter_by(style_id=style_id).count() == 0
    
    # Create database record
    new_image = StyleImage(
        style_id=style_id,
        filename=filename,
        is_primary=is_primary,
        upload_date=datetime.now()
    )
    
    db.session.add(new_image)
    # Update style's updated_at timestamp
    style.updated_at = datetime.now()
    db.session.commit()
    
    app.logger.info(f"Image uploaded: {filename} for style {style_id} ({file_size} bytes)")
    
    return jsonify({
        'success': True,
        'id': new_image.id,
        'filename': filename,
        'url': f'/static/img/{filename}',
        'is_primary': is_primary
    }), 200

@app.route('/api/style/<int:style_id>/images', methods=['GET'])
@login_required
def get_style_images(style_id):
    """Get all images for a style"""
    # Check if style exists
    style = Style.query.get_or_404(style_id)
    
    # Get all images for this style, ordered by primary first, then by upload date
    images = StyleImage.query.filter_by(
        style_id=style_id
    ).order_by(
        StyleImage.is_primary.desc(),
        StyleImage.upload_date.asc()
    ).all()
    
    # Return image data as JSON
    return jsonify([{
        'id': img.id,
        'url': f'/static/img/{img.filename}',
        'filename': img.filename,
        'is_primary': img.is_primary,
        'upload_date': img.upload_date.isoformat() if img.upload_date else None
    } for img in images]), 200

@app.route('/api/style-image/<int:image_id>', methods=['DELETE'])
@admin_required 
def delete_style_image(image_id):
    """Delete a style image"""
    # Find the image record
    img = StyleImage.query.get_or_404(image_id)
    
    # Get the style before deleting the image
    style = Style.query.get(img.style_id)
    
    # Delete physical file from disk
    filepath = os.path.join(app.config['UPLOAD_FOLDER'], img.filename)
    
    try:
        os.remove(filepath)
    except FileNotFoundError:
        pass  # File already deleted, that's fine
    except Exception as e:
        app.logger.error(f"Error deleting file: {e}")
        # Continue anyway to remove from database
    
    # Delete database record
    db.session.delete(img)
    
    # Update style's updated_at timestamp
    if style:
        style.updated_at = datetime.now()

    db.session.commit()
    
    return jsonify({'success': True}), 200


    
# SIZE RANGE ENDPOINTS
@app.route('/api/size-ranges', methods=['GET', 'POST'])
@login_required
def api_size_ranges():
    if request.method == 'GET':
        ranges = SizeRange.query.all()
        return jsonify([{
            'id': r.id,
            'name': r.name,
            'regular_sizes': r.regular_sizes,
            'extended_sizes': r.extended_sizes,
            'extended_markup_percent': r.extended_markup_percent
        } for r in ranges])
    
    elif request.method == 'POST':
        try:
            data = request.get_json()
            
            # Validation
            name, error = validate_required_string(data.get('name'), 'Size range name', max_length=50)
            if error:
                return jsonify({'success': False, 'error': error}), 400
            
            regular_sizes, error = validate_required_string(data.get('regular_sizes'), 'Regular sizes', max_length=100)
            if error:
                return jsonify({'success': False, 'error': error}), 400
            
            markup, error = validate_positive_number(data.get('extended_markup_percent', 15), 'Extended markup percent', required=False)
            if error:
                return jsonify({'success': False, 'error': error}), 400
            
            size_range = SizeRange(
                name=name,
                regular_sizes=regular_sizes,
                extended_sizes=data.get('extended_sizes', '').strip() if data.get('extended_sizes') else None,
                extended_markup_percent=markup or 15.0,
                description=data.get('description', '').strip() if data.get('description') else None
            )
            db.session.add(size_range)
            db.session.commit()
            
            return jsonify({'success': True, 'id': size_range.id})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error adding size range: {e}")
            return jsonify({'success': False, 'error': 'Failed to add size range'}), 500

@app.route('/api/size-ranges/<int:size_range_id>', methods=['GET'])
@login_required
def api_size_range_get(size_range_id):
    """Get size range details - any authenticated user"""
    size_range = SizeRange.query.get_or_404(size_range_id)
    return jsonify({
        'id': size_range.id,
        'name': size_range.name,
        'regular_sizes': size_range.regular_sizes,
        'extended_sizes': size_range.extended_sizes,
        'extended_markup_percent': size_range.extended_markup_percent,
        'description': size_range.description
    })


@app.route('/api/size-ranges/<int:size_range_id>', methods=['PUT', 'DELETE'])
@login_required
@role_required('admin')
def api_size_range_modify(size_range_id):
    """Update or delete size range - admin only"""
    size_range = SizeRange.query.get_or_404(size_range_id)
    
    if request.method == 'PUT':
        try:
            data = request.get_json()
            
            if 'name' in data:
                name, error = validate_required_string(data.get('name'), 'Size range name', max_length=50)
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                size_range.name = name
            
            if 'regular_sizes' in data:
                sizes, error = validate_required_string(data.get('regular_sizes'), 'Regular sizes', max_length=100)
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                size_range.regular_sizes = sizes
            
            if 'extended_sizes' in data:
                size_range.extended_sizes = data.get('extended_sizes', '').strip() if data.get('extended_sizes') else None
            
            if 'extended_markup_percent' in data:
                markup, error = validate_positive_number(data.get('extended_markup_percent'), 'Extended markup percent', required=False)
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                size_range.extended_markup_percent = markup or 15.0
            
            if 'description' in data:
                size_range.description = data.get('description', '').strip() if data.get('description') else None
            
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating size range: {e}")
            return jsonify({'success': False, 'error': 'Failed to update size range'}), 500
    
    elif request.method == 'DELETE':
        try:
            # Check if size range is used in any styles
            styles_using_range = Style.query.filter_by(size_range=size_range.name).all()
            if styles_using_range:
                styles_using = [{
                    'id': s.id,
                    'vendor_style': s.vendor_style,
                    'style_name': s.style_name
                } for s in styles_using_range]
                
                style_list = ', '.join([s['vendor_style'] or s['style_name'] for s in styles_using])
                return jsonify({
                    'success': False,
                    'error': f'Cannot delete: This size range is used in {len(styles_using)} style(s): {style_list}',
                    'styles_using': styles_using
                }), 400
            
            db.session.delete(size_range)
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error deleting size range: {e}")
            return jsonify({'success': False, 'error': 'Failed to delete size range'}), 500

# GLOBAL SETTINGS ENDPOINTS
@app.route('/api/global-settings', methods=['GET'])
@login_required
@role_required('admin')
def get_global_settings():
    settings = GlobalSetting.query.all()
    return jsonify([{
        'id': s.id,
        'setting_key': s.setting_key,
        'setting_value': float(s.setting_value),
        'description': s.description
    } for s in settings])

@app.route('/api/global-settings/<int:setting_id>', methods=['GET', 'PUT'])
@login_required
@role_required('admin')
def api_global_setting_detail(setting_id):
    setting = GlobalSetting.query.get_or_404(setting_id)
    
    if request.method == 'GET':
        return jsonify({
            'id': setting.id,
            'setting_key': setting.setting_key,
            'setting_value': setting.setting_value,
            'description': setting.description
        })
    
    elif request.method == 'PUT':
        try:
            data = request.get_json()
            
            if 'setting_value' in data:
                value, error = validate_positive_number(data.get('setting_value'), 'Setting value', allow_zero=True)
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                setting.setting_value = value

                # Auto-update cleaning costs when rate changes
                if setting.setting_key == 'cleaning_cost_per_minute':
                    cleaning_costs = CleaningCost.query.all()
                    for cc in cleaning_costs:
                        cc.fixed_cost = cc.avg_minutes * value
                    app.logger.info(f"Updated {len(cleaning_costs)} cleaning costs with new rate ${value}/min")
                
                # Auto-update all styles when sublimation cost changes
                if setting.setting_key == 'sublimation_cost':
                    # Find all styles that use sublimation
                    affected_style_ids = db.session.query(StyleFabric.style_id).filter(
                        StyleFabric.is_sublimation == True
                    ).distinct().all()
                    
                    for (style_id,) in affected_style_ids:
                        style = Style.query.get(style_id)
                        if style and style.base_margin_percent:
                            new_cost = style.get_total_cost()
                            margin = style.base_margin_percent / 100.0
                            if margin < 1:
                                style.suggested_price = round(new_cost / (1 - margin), 2)
                    
                    app.logger.info(f"Updated {len(affected_style_ids)} styles after sublimation cost change to ${value}")
                
                # Auto-update all styles when label cost changes
                if setting.setting_key == 'avg_label_cost':
                    all_styles = Style.query.all()
                    for style in all_styles:
                        if style.base_margin_percent:
                            new_cost = style.get_total_cost()
                            margin = style.base_margin_percent / 100.0
                            if margin < 1:
                                style.suggested_price = round(new_cost / (1 - margin), 2)
                    app.logger.info(f"Updated {len(all_styles)} styles after label cost change to ${value}")
                
                # Auto-update all styles when shipping cost changes
                if setting.setting_key == 'shipping_cost':
                    all_styles = Style.query.all()
                    for style in all_styles:
                        if style.base_margin_percent:
                            new_cost = style.get_total_cost()
                            margin = style.base_margin_percent / 100.0
                            if margin < 1:
                                style.suggested_price = round(new_cost / (1 - margin), 2)
                    app.logger.info(f"Updated {len(all_styles)} styles after shipping cost change to ${value}")

            if 'description' in data:
                setting.description = data.get('description', '').strip() if data.get('description') else None

            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating global setting: {e}")
            return jsonify({'success': False, 'error': 'Failed to update setting'}), 500


@app.route('/import-colors')
@admin_required
def import_colors():
    """One-time import of colors from Excel file"""
    import pandas as pd
    try:
        df = pd.read_excel('V100COLORS.xlsx')
        
        imported = 0
        skipped = 0
        
        for index, row in df.iterrows():
            color_name = str(row['Color']).strip().upper()
            
            if not color_name or color_name == 'NAN':
                continue
                
            # Check if color already exists
            existing = Color.query.filter(func.lower(Color.name) == color_name.lower()).first()
            if existing:
                skipped += 1
                continue
            
            # Create new color
            color = Color(name=color_name)
            db.session.add(color)
            imported += 1
        
        db.session.commit()
        
        return f"""
        <div style="max-width: 600px; margin: 50px auto; padding: 20px; font-family: Arial;">
            <h1>Color Import Complete</h1>
            <div style="background: #d4edda; padding: 15px; border-radius: 5px; margin: 20px 0;">
                <p><strong>✓ Imported: {imported} new colors</strong></p>
                <p>⊙ Skipped: {skipped} existing colors</p>
            </div>
            <a href="/master-costs" style="background: #007bff; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px;">View Master Costs</a>
            <a href="/admin-panel" style="background: #6c757d; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; margin-left: 10px;">Back to Admin</a>
        </div>
        """
    except Exception as e:
        return f"<h1>Error importing colors:</h1><p>{str(e)}</p>"
    
# COLOR ENDPOINTS
@app.route('/api/colors', methods=['GET'])
@login_required
def api_colors_get():
    """Get all colors - any authenticated user"""
    colors = Color.query.order_by(Color.name).all()
    return jsonify([{'id': c.id, 'name': c.name, 'color_code': c.color_code} for c in colors])


@app.route('/api/colors', methods=['POST'])
@login_required
@role_required('admin')
def api_colors_create():
    """Create new color - admin only"""
    try:
        data = request.get_json()
        
        # Validation
        name, error = validate_required_string(data.get('name'), 'Color name', max_length=100)
        if error:
            return jsonify({'success': False, 'error': error}), 400
        
        # Check for duplicate
        existing = Color.query.filter(func.lower(Color.name) == name.lower()).first()
        if existing:
            return jsonify({'success': False, 'error': 'Color already exists'}), 400
        
        color = Color(
            name=name,
            color_code=data.get('color_code', '').strip() if data.get('color_code') else None
        )
        db.session.add(color)
        db.session.commit()

        return jsonify({'success': True, 'id': color.id, 'name': color.name})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error adding color: {e}")
        return jsonify({'success': False, 'error': 'Failed to add color'}), 500


@app.route('/api/colors/<int:color_id>', methods=['GET'])
@login_required
def api_color_get(color_id):
    """Get color details - any authenticated user"""
    color = Color.query.get_or_404(color_id)
    return jsonify({
        'id': color.id,
        'name': color.name,
        'color_code': color.color_code
    })


@app.route('/api/colors/<int:color_id>', methods=['PUT', 'DELETE'])
@login_required
@role_required('admin')
def api_color_modify(color_id):
    """Update or delete color - admin only"""
    color = Color.query.get_or_404(color_id)
    
    if request.method == 'PUT':
        try:
            data = request.get_json()
            
            if 'name' in data:
                name, error = validate_required_string(data.get('name'), 'Color name', max_length=100)
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                color.name = name
            
            if 'color_code' in data:
                color.color_code = data.get('color_code', '').strip() if data.get('color_code') else None
            
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating color: {e}")
            return jsonify({'success': False, 'error': 'Failed to update color'}), 500
    
    elif request.method == 'DELETE':
        try:
            # Check if color is used in any styles
            style_colors = StyleColor.query.filter_by(color_id=color_id).all()
            if style_colors:
                styles_using = []
                for sc in style_colors:
                    style = Style.query.get(sc.style_id)
                    if style:
                        styles_using.append({
                            'id': style.id,
                            'vendor_style': style.vendor_style,
                            'style_name': style.style_name
                        })
                
                style_list = ', '.join([s['vendor_style'] or s['style_name'] for s in styles_using])
                return jsonify({
                    'success': False,
                    'error': f'Cannot delete: This color is used in {len(styles_using)} style(s): {style_list}',
                    'styles_using': styles_using
                }), 400
            
            db.session.delete(color)
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error deleting color: {e}")
            return jsonify({'success': False, 'error': 'Failed to delete color'}), 500


# ===== Extended-size + Range helpers (robust) =====

def _normalize_size_token(tok: str) -> str:
    """Normalize a size label to compare safely, e.g. '3x' -> '3XL'."""
    if tok is None:
        return ''
    s = tok.strip().upper().replace(' ', '')
    # Treat '3X' as '3XL'
    if s.endswith('X') and not s.endswith('XL'):
        s = s + 'L'
    return s

_ALPHA_LADDER = ["XXS","XS","S","M","L","XL","2XL","3XL","4XL","5XL","6XL"]

def _expand_alpha_range(a: str, b: str):
    A = _normalize_size_token(a); B = _normalize_size_token(b)
    try:
        ia, ib = _ALPHA_LADDER.index(A), _ALPHA_LADDER.index(B)
    except ValueError:
        return [A, B] if A != B else [A]
    return _ALPHA_LADDER[ia:ib+1] if ia <= ib else list(reversed(_ALPHA_LADDER[ib:ia+1]))

def _expand_numeric_range(a: str, b: str):
    """Expand numeric ranges with step=2."""
    sa, sb = a.strip(), b.strip()
    try:
        ia, ib = int(sa), int(sb)
    except ValueError:
        return [_normalize_size_token(a), _normalize_size_token(b)]

    if sa == "00":
        start, end = (ia, ib) if ia <= ib else (ib, ia)
        tail = [str(n) for n in range(0, end + 1, 2)]
        return ["00"] + tail

    width = max(len(sa), len(sb))
    step = 2
    if ia <= ib:
        seq = range(ia, ib + 1, step)
    else:
        seq = range(ia, ib - 1, -step)

    return [str(n).zfill(width) for n in seq]


def _expand_mixed_token(token: str):
    t = token.strip()
    if not t:
        return []
    if '-' in t:
        left, right = [x.strip() for x in t.split('-', 1)]
        if any(c.isalpha() for c in left+right):
            return _expand_alpha_range(left, right)
        return _expand_numeric_range(left, right)
    s = _normalize_size_token(t)
    return [s] if s else []

def expand_sizes_string(s: str):
    """'XS-XL, 2XL-6XL' or '00-18, 20-30' -> flat list of sizes."""
    if not s:
        return []
    out = []
    for part in s.split(','):
        out.extend(_expand_mixed_token(part))
    # de-dup preserving order
    seen, flat = set(), []
    for x in out:
        if x not in seen:
            seen.add(x); flat.append(x)
    return flat

def is_extended_size_for_range(size_label: str, size_range_obj) -> bool:
    """True if size_label is inside size_range_obj.extended_sizes (supports ranges)."""
    if size_range_obj is None:
        return False
    ext_list = expand_sizes_string(getattr(size_range_obj, 'extended_sizes', '') or '')
    if not ext_list:
        return False
    return _normalize_size_token(size_label) in ext_list


# ===== VALIDATION HELPER =====
def validate_style_for_export(style):
    """
    Validate a single style for export.
    Returns (is_valid: bool, missing: list)
    """
    missing = []
    
    # Check vendor style
    if not style.vendor_style or not style.vendor_style.strip():
        missing.append('Vendor Style')
    
    # Check style name
    if not style.style_name or not style.style_name.strip():
        missing.append('Style Name')
    
    # Check colors - STRICT (must have at least 1)
    has_colors = hasattr(style, 'colors') and style.colors and len(style.colors) > 0
    if not has_colors:
        missing.append('At least ONE Color')
    
    # Check size range - STRICT (must exist AND have sizes)
    if not style.size_range or not style.size_range.strip():
        missing.append('Size Range')
    else:
        # Verify size range actually has sizes defined
        from models import SizeRange
        size_range_obj = SizeRange.query.filter_by(name=style.size_range).first()
        if size_range_obj:
            regular_sizes = expand_sizes_string(getattr(size_range_obj, 'regular_sizes', '') or '')
            extended_sizes = expand_sizes_string(getattr(size_range_obj, 'extended_sizes', '') or '')
            all_sizes = regular_sizes + [s for s in extended_sizes if s not in regular_sizes]
            
            if not all_sizes:
                missing.append('Size Range has NO sizes defined')
        else:
            missing.append('Size Range does not exist in system')
    
    return (len(missing) == 0, missing)


# ===== SINGLE STYLE EXPORT =====
@app.route('/export-sap-single-style', methods=['POST'])
@login_required
def export_sap_single_style():
    """Export a single style in SAP B1 format - STRICT VALIDATION"""
    try:
        vendor_style = request.form.get('vendor_style')
        if not vendor_style:
            return "No style specified", 400

        style = Style.query.filter_by(vendor_style=vendor_style).first()
        if not style:
            return "Style not found", 404

        # ========================================
        # STRICT VALIDATION
        # ========================================
        is_valid, missing = validate_style_for_export(style)
        
        if not is_valid:
            error_html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Export Validation Error</title>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        padding: 40px;
                        background: #f5f5f5;
                    }}
                    .container {{
                        max-width: 700px;
                        margin: 0 auto;
                        background: white;
                        padding: 30px;
                        border-radius: 8px;
                        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                    }}
                    h1 {{
                        color: #dc3545;
                        margin-bottom: 20px;
                    }}
                    .style-info {{
                        background: #f8f9fa;
                        padding: 15px;
                        border-radius: 4px;
                        margin-bottom: 20px;
                    }}
                    .error-box {{
                        background: #fff3cd;
                        border-left: 4px solid #ffc107;
                        padding: 15px;
                        margin: 20px 0;
                    }}
                    .missing-item {{
                        color: #856404;
                        margin: 5px 0;
                        padding-left: 20px;
                    }}
                    .missing-item::before {{
                        content: "⚠️ ";
                    }}
                    .back-btn {{
                        display: inline-block;
                        margin-top: 20px;
                        padding: 10px 20px;
                        background: #007bff;
                        color: white;
                        text-decoration: none;
                        border-radius: 4px;
                    }}
                    .back-btn:hover {{
                        background: #0056b3;
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <h1>❌ Cannot Export</h1>
                    
                    <div class="style-info">
                        <strong>Style:</strong> {html.escape(style.vendor_style or '')}<br>
                        <strong>Name:</strong> {html.escape(style.style_name or 'N/A')}
                    </div>
                    
                    <div class="error-box">
                        <strong>Missing Required Fields:</strong>
                        {''.join(f'<div class="missing-item">{html.escape(item)}</div>' for item in missing)}
                    </div>
                    
                    <p><strong>All fields below are REQUIRED for export:</strong></p>
                    <ul>
                        <li>Vendor Style</li>
                        <li>Style Name</li>
                        <li>At least ONE Color</li>
                        <li>Size Range (with sizes defined)</li>
                    </ul>
                    
                    <a href="/style/new?vendor_style={html.escape(vendor_style or '')}" class="back-btn">← Edit This Style</a>
                </div>
            </body>
            </html>
            """
            return error_html, 400

        # ========================================
        # EXPORT - NO FALLBACKS
        # ========================================
        output = StringIO()
        writer = csv.writer(output)

        headers = ['Code', 'Name', 'U_COLOR', 'U_SIZE', 'U_VARIABLE',
                   'U_PRICE', 'U_SHIP_COST', 'U_STYLE', 'U_CardCode', 'U_PROD_NAME']
        writer.writerow(headers)
        writer.writerow(headers)

        base_cost = style.get_total_cost()
        from models import SizeRange
        size_range_obj = SizeRange.query.filter_by(name=style.size_range).first()

        extended_pct = (size_range_obj.extended_markup_percent
                        if (size_range_obj and size_range_obj.extended_markup_percent is not None)
                        else 15.0)
        extended_mult = 1.0 + (float(extended_pct) / 100.0)

        regular_list = expand_sizes_string(getattr(size_range_obj, 'regular_sizes', '') or '')
        extended_list = expand_sizes_string(getattr(size_range_obj, 'extended_sizes', '') or '')
        all_sizes = regular_list + [s for s in extended_list if s not in regular_list]
        
        # NO FALLBACK - validation ensures sizes exist
        colors = [sc.color.name.upper() for sc in style.colors]
        
        # Get variables - DEFAULT exports as empty string, others export as-is
        variables = []
        if hasattr(style, 'style_variables') and style.style_variables:
            for sv in style.style_variables:
                var_name = sv.variable.name.upper()
                if var_name == 'DEFAULT':
                    variables.append('')  # DEFAULT = empty in export
                else:
                    variables.append(var_name)  # REGULAR, TALL, etc. = as-is

        if not variables:
            variables = ['']  # Fallback if no variables at all
        
        vendor_code = 'V100'

        u_style = style.vendor_style.replace('-', '')
        shipping_cost = style.shipping_cost if hasattr(style, 'shipping_cost') else 0.00

        for color in colors:
                for size in all_sizes:
                    price = round(base_cost * extended_mult, 2) if is_extended_size_for_range(size, size_range_obj) else round(base_cost, 2)
                    for variable in variables:
                        writer.writerow(['', '', color, size, variable, price, shipping_cost, u_style, vendor_code, style.style_name])

        output.seek(0)
        filename = f"SAP_{style.vendor_style}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return Response(output.getvalue(), mimetype='text/csv',
                        headers={'Content-Disposition': f'attachment; filename={filename}'})

    except Exception as e:
        import traceback
        app.logger.error(traceback.format_exc())
        return f"Error exporting: {html.escape(str(e))}", 500


# ===== BULK EXPORT =====
@app.route('/export-sap-format', methods=['POST'])
@login_required
def export_sap_format():
    """Export selected styles in SAP B1 format - STRICT VALIDATION"""
    try:
        import json
        style_ids = json.loads(request.form.get('style_ids', '[]'))
        include_empty_vars = request.form.get('include_empty_vars', '0') == '1'
        if not style_ids:
            return "No styles selected", 400

        styles = Style.query.filter(Style.id.in_(style_ids)).all()
        
        # ========================================
        # STRICT VALIDATION - ALL STYLES
        # ========================================
        invalid_styles = []
        
        for style in styles:
            is_valid, missing = validate_style_for_export(style)
            if not is_valid:
                invalid_styles.append({
                    'vendor_style': style.vendor_style or 'UNKNOWN',
                    'style_name': style.style_name or 'UNKNOWN',
                    'missing': missing
                })
        
        # If ANY style is invalid, block export and show errors
        if invalid_styles:
            error_html = f"""
            <!DOCTYPE html>
            <html>
            <head>
                <title>Export Validation Error</title>
                <style>
                    body {{
                        font-family: Arial, sans-serif;
                        padding: 40px;
                        background: #f5f5f5;
                    }}
                    .container {{
                        max-width: 900px;
                        margin: 0 auto;
                        background: white;
                        padding: 30px;
                        border-radius: 8px;
                        box-shadow: 0 2px 10px rgba(0,0,0,0.1);
                    }}
                    h1 {{
                        color: #dc3545;
                        margin-bottom: 20px;
                    }}
                    .summary {{
                        background: #fff3cd;
                        border: 2px solid #ffc107;
                        padding: 15px;
                        margin-bottom: 30px;
                        border-radius: 4px;
                        font-weight: bold;
                        text-align: center;
                    }}
                    .error-list {{
                        background: #f8f9fa;
                        padding: 20px;
                        border-radius: 4px;
                    }}
                    .style-error {{
                        margin-bottom: 20px;
                        padding: 15px;
                        border-left: 4px solid #dc3545;
                        background: white;
                    }}
                    .vendor-style {{
                        font-weight: bold;
                        color: #333;
                        font-size: 1.1em;
                        margin-bottom: 5px;
                    }}
                    .style-name {{
                        color: #666;
                        font-size: 0.9em;
                        margin-bottom: 10px;
                    }}
                    .missing-item {{
                        color: #856404;
                        margin: 3px 0;
                        padding-left: 20px;
                    }}
                    .missing-item::before {{
                        content: "⚠️ ";
                    }}
                    .requirements {{
                        background: #e7f3ff;
                        border-left: 4px solid #007bff;
                        padding: 15px;
                        margin: 20px 0;
                    }}
                    .back-btn {{
                        display: inline-block;
                        margin-top: 20px;
                        padding: 12px 24px;
                        background: #007bff;
                        color: white;
                        text-decoration: none;
                        border-radius: 4px;
                        font-weight: bold;
                    }}
                    .back-btn:hover {{
                        background: #0056b3;
                    }}
                </style>
            </head>
            <body>
                <div class="container">
                    <h1>❌ Export Blocked - Validation Failed</h1>
                    
                    <div class="summary">
                        {len(invalid_styles)} of {len(styles)} selected style(s) are missing required fields
                    </div>
                    
                    <div class="error-list">
            """
            
            for inv in invalid_styles:
                error_html += f"""
                        <div class="style-error">
                            <div class="vendor-style">{html.escape(inv['vendor_style'])}</div>
                            <div class="style-name">{html.escape(inv['style_name'])}</div>
                            {''.join(f'<div class="missing-item">{html.escape(item)}</div>' for item in inv['missing'])}
                        </div>
                """
            
            error_html += """
                    </div>
                    
                    <div class="requirements">
                        <strong>🔒 All fields below are REQUIRED for export:</strong>
                        <ul>
                            <li><strong>Vendor Style</strong> - Must be filled</li>
                            <li><strong>Style Name</strong> - Must be filled</li>
                            <li><strong>At least ONE Color</strong> - Add colors to the style</li>
                            <li><strong>Size Range</strong> - Must be selected with sizes defined</li>
                        </ul>
                    </div>
                    
                    <p>Please complete ALL required fields for ALL selected styles before exporting.</p>
                    
                    <a href="/view-all-styles" class="back-btn">← Back to All Styles</a>
                </div>
            </body>
            </html>
            """
            
            return error_html, 400
        
        # ========================================
        # ALL VALIDATED - EXPORT (NO FALLBACKS)
        # ========================================
        output = StringIO()
        writer = csv.writer(output)

        headers = ['Code', 'Name', 'U_COLOR', 'U_SIZE', 'U_VARIABLE',
                   'U_PRICE', 'U_SHIP_COST', 'U_STYLE', 'U_CardCode', 'U_PROD_NAME']
        writer.writerow(headers)
        writer.writerow(headers)

        for style in styles:
            base_cost = style.get_total_cost()
            from models import SizeRange
            size_range_obj = SizeRange.query.filter_by(name=style.size_range).first()

            extended_pct = (size_range_obj.extended_markup_percent
                            if (size_range_obj and size_range_obj.extended_markup_percent is not None)
                            else 15.0)
            extended_mult = 1.0 + (float(extended_pct) / 100.0)

            regular_list = expand_sizes_string(getattr(size_range_obj, 'regular_sizes', '') or '')
            extended_list = expand_sizes_string(getattr(size_range_obj, 'extended_sizes', '') or '')
            all_sizes = regular_list + [s for s in extended_list if s not in regular_list]

            # NO FALLBACKS - validation ensures these exist
            colors = [sc.color.name.upper() for sc in style.colors]
            
            variables = [sv.variable.name.upper() for sv in style.style_variables] if hasattr(style, 'style_variables') else []
            vendor_code = 'V100'

            u_style = style.vendor_style.replace('-', '')
            shipping_cost = style.shipping_cost if hasattr(style, 'shipping_cost') else 0.00

            for color in colors:
                for size in all_sizes:
                    price = round(base_cost * extended_mult, 2) if is_extended_size_for_range(size, size_range_obj) else round(base_cost, 2)
                    if variables:
                        for variable in variables:
                            writer.writerow(['', '', color, size, variable, price, shipping_cost, u_style, vendor_code, style.style_name])
                        # Only add empty variable row if include_empty_vars is True
                        if include_empty_vars:
                            writer.writerow(['', '', color, size, '', price, shipping_cost, u_style, vendor_code, style.style_name])
                    else:
                        writer.writerow(['', '', color, size, '', price, shipping_cost, u_style, vendor_code, style.style_name])

        output.seek(0)
        filename = f"SAP_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        return Response(output.getvalue(), mimetype='text/csv',
                        headers={'Content-Disposition': f'attachment; filename={filename}'})

    except Exception as e:
        import traceback
        app.logger.error(traceback.format_exc())
        return f"Error exporting: {html.escape(str(e))}", 500


@app.route('/view-all-styles')
@role_required('admin', 'user')
def view_all_styles():
    """View all styles with filters - RBAC enabled (Admin and User only)"""
    # Eager load to avoid N+1 queries
    styles = Style.query.options(
        joinedload(Style.style_fabrics).joinedload(StyleFabric.fabric),
        joinedload(Style.style_notions).joinedload(StyleNotion.notion),
        joinedload(Style.style_labor).joinedload(StyleLabor.labor_operation)
    ).all()

    # Cache label cost
    label_setting = GlobalSetting.query.filter_by(setting_key='avg_label_cost').first()
    label_cost = label_setting.setting_value if label_setting else 0.20

    # Calculate stats
    total_styles = len(styles)
    total_value = sum(s.get_total_fabric_cost() + s.get_total_notion_cost() + s.get_total_labor_cost() + label_cost for s in styles)
    avg_cost = total_value / total_styles if total_styles > 0 else 0
    
    # Get user permissions for frontend
    permissions = get_user_permissions()
    
    return render_template('view_all_styles.html', 
                         styles=styles,
                         total_styles=total_styles,
                         total_value=total_value,
                         avg_cost=avg_cost,
                         permissions=permissions,
                         current_user=current_user)

    
@app.route('/api/style/delete/<int:style_id>', methods=['DELETE'])
@limiter.limit("10 per minute")
@admin_required 
def delete_style(style_id):
    """Delete a style and all its relationships including physical image files"""
    try:
        style = Style.query.get_or_404(style_id)
        
        # Save style info for audit log BEFORE deleting
        style_info = {
            "vendor_style": style.vendor_style,
            "style_name": style.style_name,
            "total_cost": str(style.get_total_cost())
        }
        
        # Delete physical image files FIRST (before DB records)
        images = StyleImage.query.filter_by(style_id=style_id).all()
        for img in images:
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], img.filename)
            try:
                os.remove(filepath)
                app.logger.info(f"Deleted image file: {filepath}")
            except FileNotFoundError:
                pass  # File already deleted, that's fine
            except Exception as e:
                app.logger.warning(f"Could not delete image file {filepath}: {e}")
                # Continue anyway - don't fail the whole delete
        
        # Now delete database records
        StyleImage.query.filter_by(style_id=style_id).delete()
        StyleFabric.query.filter_by(style_id=style_id).delete()
        StyleNotion.query.filter_by(style_id=style_id).delete()
        StyleLabor.query.filter_by(style_id=style_id).delete()
        StyleColor.query.filter_by(style_id=style_id).delete()
        StyleVariable.query.filter_by(style_id=style_id).delete()
        
        # Delete the style itself
        db.session.delete(style)
        db.session.commit()
        
        # Log audit after successful deletion
        try:
            log_audit(
                action="DELETE",
                item_type="style",
                item_id=style_id,
                item_name=style_info["vendor_style"],
                old_values=style_info,
                new_values=None,
                details=f"Deleted style: {style_info['style_name']}"
            )
        except Exception as e:
            app.logger.error(f"Failed to log audit for style deletion: {e}")
        
        app.logger.info(f"Style {style_id} deleted successfully with {len(images)} image(s)")
        
        return jsonify({"success": True, "message": "Style deleted successfully"}), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        error_details = traceback.format_exc()
        app.logger.error(f"Delete error: {error_details}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/style/duplicate/<int:style_id>', methods=['POST'])
@admin_required 
def duplicate_style(style_id):
    """Duplicate a style with all its components"""
    try:
        original = Style.query.get_or_404(style_id)
        
        # Generate unique vendor_style for the copy
        base_vendor_style = original.vendor_style if original.vendor_style else "COPY"
        new_vendor_style = f"{base_vendor_style}-COPY"
        
        # Check if this vendor_style already exists, if so add a number
        counter = 1
        while Style.query.filter_by(vendor_style=new_vendor_style).first():
            new_vendor_style = f"{base_vendor_style}-COPY{counter}"
            counter += 1
        
        # Create new style
        new_style = Style(
            style_name=f"{original.style_name} (Copy)",
            vendor_style=new_vendor_style,
            base_item_number=original.base_item_number,
            variant_code=original.variant_code,
            gender=original.gender,
            garment_type=original.garment_type,
            size_range=original.size_range,
            notes=original.notes,
            base_margin_percent=original.base_margin_percent,
            suggested_price=original.suggested_price
        )
        db.session.add(new_style)
        db.session.flush()
        
        # Copy fabrics - Query directly from StyleFabric table
        style_fabrics = StyleFabric.query.filter_by(style_id=original.id).all()
        for sf in style_fabrics:
            new_sf = StyleFabric(
                style_id=new_style.id,
                fabric_id=sf.fabric_id,
                yards_required=sf.yards_required,
                is_primary=sf.is_primary,
                is_sublimation=sf.is_sublimation
            )
            db.session.add(new_sf)
        
        # Copy notions - Query directly from StyleNotion table
        style_notions = StyleNotion.query.filter_by(style_id=original.id).all()
        for sn in style_notions:
            new_sn = StyleNotion(
                style_id=new_style.id,
                notion_id=sn.notion_id,
                quantity_required=sn.quantity_required
            )
            db.session.add(new_sn)
        
        # Copy labor - Query directly from StyleLabor table
        style_labor = StyleLabor.query.filter_by(style_id=original.id).all()
        for sl in style_labor:
            new_sl = StyleLabor(
                style_id=new_style.id,
                labor_operation_id=sl.labor_operation_id,
                time_hours=sl.time_hours,
                quantity=sl.quantity
            )
            db.session.add(new_sl)
        
        # Copy colors - Query directly from StyleColor table
        style_colors = StyleColor.query.filter_by(style_id=original.id).all()
        for sc in style_colors:
            new_sc = StyleColor(
                style_id=new_style.id,
                color_id=sc.color_id
            )
            db.session.add(new_sc)
        
        # Copy variables - Query directly from StyleVariable table
        style_variables = StyleVariable.query.filter_by(style_id=original.id).all()
        for sv in style_variables:
            new_sv = StyleVariable(
                style_id=new_style.id,
                variable_id=sv.variable_id
            )
            db.session.add(new_sv)
        
        db.session.commit()
        
        return jsonify({
            "success": True, 
            "new_style_id": new_style.id, 
            "new_vendor_style": new_vendor_style,
            "message": f"Style duplicated as '{new_vendor_style}'"
        }), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        error_details = traceback.format_exc()
        app.logger.error(f"Duplicate error: {error_details}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/styles/bulk-delete', methods=['POST'])
@limiter.limit("3 per minute")
@admin_required 
def bulk_delete_styles():
    """Delete multiple styles including their physical image files"""
    try:
        data = request.get_json()
        style_ids = data.get('style_ids', [])
        
        if not style_ids:
            return jsonify({"success": False, "error": "No styles selected"}), 400
        
        total_images_deleted = 0
        
        for style_id in style_ids:
            # ✅ NEW: Delete physical image files FIRST
            images = StyleImage.query.filter_by(style_id=style_id).all()
            for img in images:
                filepath = os.path.join(app.config['UPLOAD_FOLDER'], img.filename)
                try:
                    os.remove(filepath)
                    total_images_deleted += 1
                except FileNotFoundError:
                    pass  # File already deleted, that's fine
                except Exception as e:
                    app.logger.warning(f"Could not delete image file {filepath}: {e}")
            
            # Delete database records
            StyleImage.query.filter_by(style_id=style_id).delete()
            StyleFabric.query.filter_by(style_id=style_id).delete()
            StyleNotion.query.filter_by(style_id=style_id).delete()
            StyleLabor.query.filter_by(style_id=style_id).delete()
            StyleColor.query.filter_by(style_id=style_id).delete()
            StyleVariable.query.filter_by(style_id=style_id).delete()
            
            # Delete the style itself
            style = Style.query.get(style_id)
            if style:
                db.session.delete(style)
        
        db.session.commit()
        
        app.logger.info(f"Bulk delete: {len(style_ids)} style(s) and {total_images_deleted} image file(s) deleted")
        
        return jsonify({
            "success": True, 
            "message": f"{len(style_ids)} style(s) deleted successfully"
        }), 200
        
    except Exception as e:
        db.session.rollback()
        import traceback
        error_details = traceback.format_exc()
        app.logger.error(f"Bulk delete error: {error_details}")
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/api/style/load-for-duplicate/<int:style_id>')
@login_required
def load_style_for_duplicate(style_id):
    """Load a style's data for duplication"""
    try:
        style = Style.query.get_or_404(style_id)
        
        # Get fabrics
        fabrics = []
        for sf in StyleFabric.query.filter_by(style_id=style.id).all():
            fabric = Fabric.query.get(sf.fabric_id)
            if fabric:
                vendor_name = ""
                if fabric.fabric_vendor_id:
                    vendor = FabricVendor.query.get(fabric.fabric_vendor_id)
                    if vendor:
                        vendor_name = vendor.name
                
                fabrics.append({
                    "name": fabric.name,
                    "vendor": vendor_name,
                    "yards": sf.yards_required,
                    "cost_per_yard": fabric.cost_per_yard,
                    "primary": sf.is_primary,
                    "sublimation": sf.is_sublimation
                })
        
        # Get notions
        notions = []
        for sn in StyleNotion.query.filter_by(style_id=style.id).all():
            notion = Notion.query.get(sn.notion_id)
            if notion:
                vendor_name = ""
                if notion.notion_vendor_id:
                    vendor = NotionVendor.query.get(notion.notion_vendor_id)
                    if vendor:
                        vendor_name = vendor.name
                
                notions.append({
                    "name": notion.name,
                    "vendor": vendor_name,
                    "qty": float(sn.quantity_required) if sn.quantity_required else 0,
                    "cost_per_unit": notion.cost_per_unit
                })
        
        # Get labor
        labor = []
        for sl in StyleLabor.query.filter_by(style_id=style.id).all():
            op = LaborOperation.query.get(sl.labor_operation_id)
            if op:
                labor.append({
                    "name": op.name,
                    "qty_or_hours": sl.time_hours if op.cost_type == 'hourly' else sl.quantity
                })
        
        # Get colors
        colors = []
        for sc in StyleColor.query.filter_by(style_id=style.id).all():
            color = Color.query.get(sc.color_id)
            if color:
                colors.append({
                    "color_id": color.id,
                    "name": color.name
                })
        
        # Get variables
        variables = []
        for sv in StyleVariable.query.filter_by(style_id=style.id).all():
            variable = Variable.query.get(sv.variable_id)
            if variable:
                variables.append({
                    "variable_id": variable.id,
                    "name": variable.name
                })
        
        # Get cleaning cost
        cleaning_data = None
        if style.garment_type:
            cleaning_op = LaborOperation.query.filter_by(name='Cleaning & Ironing').first()
            if cleaning_op:
                cleaning_cost_record = CleaningCost.query.filter_by(garment_type=style.garment_type).first()
                if cleaning_cost_record:
                    cleaning_data = {
                        "garment_type": style.garment_type,
                        "cost": cleaning_cost_record.fixed_cost
                    }

         # Get label and shipping costs from global settings
        label_setting = GlobalSetting.query.filter_by(setting_key='avg_label_cost').first()
        shipping_setting = GlobalSetting.query.filter_by(setting_key='shipping_cost').first()
        label_cost_value = label_setting.setting_value if label_setting else 0.20
        shipping_cost_value = shipping_setting.setting_value if shipping_setting else 0.00

        return jsonify({
            "success": True,
            "style": {
                "vendor_style": style.vendor_style + "-COPY",  # Suggest a new vendor style
                "style_name": style.style_name + " (Copy)",     # Suggest a new style name
                "base_item_number": style.base_item_number,
                "variant_code": style.variant_code,
                "gender": style.gender,
                "garment_type": style.garment_type,
                "size_range": style.size_range,
                "margin": style.base_margin_percent,
                "suggested_price": style.suggested_price,
                "notes": style.notes,
                "label_cost": label_cost_value,
                "shipping_cost": shipping_cost_value,
                "original_vendor_style": style.vendor_style,  # Track original
                "original_style_name": style.style_name        # Track original
            },
            "fabrics": fabrics,
            "notions": notions,
            "labor": labor,
            "cleaning": cleaning_data,
            "colors": colors,
            "variables": variables
        }), 200
    

        
    except Exception as e:
        return jsonify({"success": False, "error": str(e)}), 500
    
@app.route('/api/style/check-vendor-style')
@login_required
def check_vendor_style():
    """Check if a vendor_style already exists"""
    vendor_style = request.args.get('vendor_style', '').strip()
    
    if not vendor_style:
        return jsonify({"exists": False})
    
    exists = Style.query.filter_by(vendor_style=vendor_style).first() is not None
    
    return jsonify({"exists": exists})


# ===== EDITABLE MASTER COSTS WITH VENDOR MANAGEMENT =====
 
@app.route('/api/style/<int:style_id>/favorite', methods=['POST'])
@admin_required
def toggle_favorite(style_id):
    """Toggle favorite status"""
    try:
        style = Style.query.get_or_404(style_id)
        data = request.get_json()
        style.is_favorite = data.get('is_favorite', False)
        db.session.commit()
        return jsonify({"success": True}), 200
    except Exception as e:
        db.session.rollback()
        return jsonify({"success": False, "error": str(e)}), 500

@app.route('/master-costs')
@login_required
def master_costs():
    """Display editable master cost lists"""
    fabrics = Fabric.query.order_by(Fabric.name).all()
    notions = Notion.query.order_by(Notion.name).all()
    labor_ops = LaborOperation.query.order_by(LaborOperation.name).all()
    cleaning_costs = CleaningCost.query.order_by(CleaningCost.garment_type).all()
    fabric_vendors = FabricVendor.query.order_by(FabricVendor.name).all()
    notion_vendors = NotionVendor.query.order_by(NotionVendor.name).all()
    colors = Color.query.order_by(Color.name).all()
    variables = Variable.query.order_by(Variable.name).all()
    size_ranges = SizeRange.query.order_by(SizeRange.name).all()
    
    global_settings = GlobalSetting.query.all()
    
    # Get cleaning cost per minute for the modal
    cleaning_rate_setting = GlobalSetting.query.filter_by(setting_key='cleaning_cost_per_minute').first()
    cleaning_cost_per_minute = cleaning_rate_setting.setting_value if cleaning_rate_setting else 0.32

    return render_template('master_costs.html',
                         fabrics=fabrics,
                         notions=notions,
                         labor_costs=labor_ops, 
                         cleaning_costs=cleaning_costs,
                         fabric_vendors=fabric_vendors,
                         notion_vendors=notion_vendors,
                         colors=colors,
                         variables=variables,
                         size_ranges=size_ranges,
                         global_settings=global_settings,
                         cleaning_cost_per_minute=cleaning_cost_per_minute)

# ===== API ENDPOINTS FOR MASTER COSTS =====

# =====================================================
# UPDATED CRUD ROUTES WITH AUDIT LOGGING
# Replace your existing routes in app.py with these
# =====================================================

# FABRIC VENDOR ENDPOINTS
@app.route('/api/fabric-vendors/<int:vendor_id>', methods=['GET', 'PUT', 'DELETE'])
@role_required('admin')
def api_fabric_vendor_detail(vendor_id):
    print(f">>> ROUTE HIT! Method: {request.method}, ID: {vendor_id}")
    vendor = FabricVendor.query.get_or_404(vendor_id)
    
    if request.method == 'GET':
        return jsonify({
            'id': vendor.id,
            'name': vendor.name,
            'vendor_code': vendor.vendor_code,
            'f_ship_cost': vendor.f_ship_cost or 0.0
        })
    
    elif request.method == 'PUT':
        try:
            print(">>> FABRIC VENDOR PUT ROUTE HIT") 
            data = request.get_json()
            
            # Capture old values BEFORE update
            old_values = {'name': vendor.name, 'code': vendor.vendor_code}
            
            if 'name' in data:
                name, error = validate_required_string(data.get('name'), 'Vendor name', max_length=100)
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                vendor.name = name
            
            if 'vendor_code' in data:
                vendor.vendor_code = data.get('vendor_code', '').strip() if data.get('vendor_code') else None

            if 'f_ship_cost' in data:
                f_ship_cost = float(data.get('f_ship_cost') or 0.0)
                if f_ship_cost < 0:
                    return jsonify({'error': 'Shipping cost cannot be negative'}), 400
                vendor.f_ship_cost = f_ship_cost
            
            db.session.commit()

            # Auto-update all styles using this vendor's fabrics
            if 'f_ship_cost' in data:
                # Find all fabrics from this vendor
                fabrics = Fabric.query.filter_by(fabric_vendor_id=vendor.id).all()
                fabric_ids = [f.id for f in fabrics]
                
                if fabric_ids:
                    # Find all styles using these fabrics
                    affected_style_ids = db.session.query(StyleFabric.style_id).filter(
                        StyleFabric.fabric_id.in_(fabric_ids)
                    ).distinct().all()
                    
                    for (style_id,) in affected_style_ids:
                        style = Style.query.get(style_id)
                        if style and style.base_margin_percent:
                            # Recalculate suggested_price based on new cost and stored margin
                            new_cost = style.get_total_cost()
                            margin = style.base_margin_percent / 100.0
                            if margin < 1:  # Prevent division by zero
                                style.suggested_price = round(new_cost / (1 - margin), 2)
                    
                    db.session.commit()
                    app.logger.info(f"Auto-updated {len(affected_style_ids)} styles after f_ship_cost change for vendor {vendor.name}")
 
            # Log the update
            log_audit(
                action='UPDATE',
                item_type='fabric_vendor',
                item_id=vendor.id,
                item_name=vendor.name,
                old_values=old_values,
                new_values={'name': vendor.name, 'code': vendor.vendor_code}
            )
            
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating fabric vendor: {e}")
            return jsonify({'success': False, 'error': 'Failed to update fabric vendor'}), 500
    
    elif request.method == 'DELETE':
        try:
            # Capture info BEFORE delete
            vendor_name = vendor.name
            vendor_id_val = vendor.id
            old_values = {'name': vendor.name, 'code': vendor.vendor_code}
            
            db.session.delete(vendor)
            db.session.commit()
            
            # Log the delete
            log_audit(
                action='DELETE',
                item_type='fabric_vendor',
                item_id=vendor_id_val,
                item_name=vendor_name,
                old_values=old_values
            )
            
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error deleting fabric vendor: {e}")
            return jsonify({'success': False, 'error': 'Failed to delete fabric vendor'}), 500

@app.route('/api/fabric-vendors', methods=['POST'])
@role_required('admin')
def api_add_fabric_vendor():
    try:
        data = request.get_json()
        
        # Validation
        name, error = validate_required_string(data.get('name'), 'Vendor name', max_length=100)
        if error:
            return jsonify({'success': False, 'error': error}), 400
        
        vendor = FabricVendor(
            name=name,
            vendor_code=data.get('vendor_code', '').strip() if data.get('vendor_code') else None,
            f_ship_cost=float(data.get('f_ship_cost') or 0.0)
        )
        db.session.add(vendor)
        db.session.commit()
        
        # Log the create
        log_audit(
            action='CREATE',
            item_type='fabric_vendor',
            item_id=vendor.id,
            item_name=vendor.name,
            new_values={'name': vendor.name, 'code': vendor.vendor_code}
        )
        
        return jsonify({'success': True, 'id': vendor.id})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error adding fabric vendor: {e}")
        return jsonify({'success': False, 'error': 'Failed to add fabric vendor'}), 500


# NOTION VENDOR ENDPOINTS
@app.route('/api/notion-vendors/<int:vendor_id>', methods=['GET', 'PUT', 'DELETE'])
@role_required('admin')
def api_notion_vendor_detail(vendor_id):
    vendor = NotionVendor.query.get_or_404(vendor_id)
    
    if request.method == 'GET':
        return jsonify({
            'id': vendor.id,
            'name': vendor.name,
            'vendor_code': vendor.vendor_code
        })
    
    elif request.method == 'PUT':
        try:
            data = request.get_json()
            
            # Capture old values BEFORE update
            old_values = {'name': vendor.name, 'code': vendor.vendor_code}
            
            if 'name' in data:
                name, error = validate_required_string(data.get('name'), 'Vendor name', max_length=100)
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                vendor.name = name
            
            if 'vendor_code' in data:
                vendor.vendor_code = data.get('vendor_code', '').strip() if data.get('vendor_code') else None
            
            db.session.commit()
            
            # Log the update
            log_audit(
                action='UPDATE',
                item_type='notion_vendor',
                item_id=vendor.id,
                item_name=vendor.name,
                old_values=old_values,
                new_values={'name': vendor.name, 'code': vendor.vendor_code}
            )
            
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating notion vendor: {e}")
            return jsonify({'success': False, 'error': 'Failed to update notion vendor'}), 500
    
    elif request.method == 'DELETE':
        try:
            # Capture info BEFORE delete
            vendor_name = vendor.name
            vendor_id_val = vendor.id
            old_values = {'name': vendor.name, 'code': vendor.vendor_code}
            
            db.session.delete(vendor)
            db.session.commit()
            
            # Log the delete
            log_audit(
                action='DELETE',
                item_type='notion_vendor',
                item_id=vendor_id_val,
                item_name=vendor_name,
                old_values=old_values
            )
            
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error deleting notion vendor: {e}")
            return jsonify({'success': False, 'error': 'Failed to delete notion vendor'}), 500


@app.route('/api/notion-vendors', methods=['POST'])
@role_required('admin')
def api_add_notion_vendor():
    try:
        data = request.get_json()
        
        # Validation
        name, error = validate_required_string(data.get('name'), 'Vendor name', max_length=100)
        if error:
            return jsonify({'success': False, 'error': error}), 400
        
        vendor = NotionVendor(
            name=name,
            vendor_code=data.get('vendor_code', '').strip() if data.get('vendor_code') else None
        )
        db.session.add(vendor)
        db.session.commit()
        
        # Log the create
        log_audit(
            action='CREATE',
            item_type='notion_vendor',
            item_id=vendor.id,
            item_name=vendor.name,
            new_values={'name': vendor.name, 'code': vendor.vendor_code}
        )
        
        return jsonify({'success': True, 'id': vendor.id})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error adding notion vendor: {e}")
        return jsonify({'success': False, 'error': 'Failed to add notion vendor'}), 500


# FABRIC ENDPOINTS   
@app.route('/api/fabrics/<int:fabric_id>', methods=['GET', 'PUT', 'DELETE'])
@role_required('admin')
def api_fabric_detail(fabric_id):
    fabric = Fabric.query.get_or_404(fabric_id)
    
    if request.method == 'GET':
        return jsonify({
            'id': fabric.id,
            'name': fabric.name,
            'cost_per_yard': fabric.cost_per_yard,
            'fabric_vendor_id': fabric.fabric_vendor_id,
            'fabric_code': fabric.fabric_code,
            'color': fabric.color
        })
    
    elif request.method == 'PUT':
        try:
            data = request.get_json()
            
            # Capture old values BEFORE update
            old_values = {
                'name': fabric.name,
                'code': fabric.fabric_code,
                'cost_per_yard': float(fabric.cost_per_yard) if fabric.cost_per_yard else None,
                'vendor': fabric.fabric_vendor.name if fabric.fabric_vendor else None
            }
            
            if 'name' in data:
                name, error = validate_required_string(data.get('name'), 'Fabric name', max_length=100)
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                fabric.name = name
            
            if 'cost_per_yard' in data:
                cost, error = validate_positive_number(data.get('cost_per_yard'), 'Cost per yard')
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                old_cost = fabric.cost_per_yard
                fabric.cost_per_yard = cost
                
                # Auto-update styles if cost changed
                if old_cost != cost:
                    affected_style_ids = db.session.query(StyleFabric.style_id).filter(
                        StyleFabric.fabric_id == fabric.id
                    ).distinct().all()
                    
                    for (style_id,) in affected_style_ids:
                        style = Style.query.get(style_id)
                        if style and style.base_margin_percent:
                            new_cost = style.get_total_cost()
                            margin = style.base_margin_percent / 100.0
                            if margin < 1:
                                style.suggested_price = round(new_cost / (1 - margin), 2)
            
            if 'fabric_vendor_id' in data:
                fabric.fabric_vendor_id = data.get('fabric_vendor_id')

            
            if 'fabric_code' in data:
                new_code = data.get('fabric_code', '').strip().upper() if data.get('fabric_code') else None
                if not new_code:
                    return jsonify({'success': False, 'error': 'Fabric code is required'}), 400
                if not new_code.startswith('T'):
                    return jsonify({'success': False, 'error': 'Fabric code must start with T (e.g., T1, T2, T3)'}), 400
                # Check if new code already exists (but not for this fabric)
                existing = Fabric.query.filter(Fabric.fabric_code == new_code, Fabric.id != fabric_id).first()
                if existing:
                    return jsonify({'success': False, 'error': f'Fabric code "{new_code}" already exists'}), 400
                fabric.fabric_code = new_code
            
            if 'color' in data:
                fabric.color = data.get('color', '').strip() if data.get('color') else None
            
            db.session.commit()
            
            # Count affected styles
            affected_count = StyleFabric.query.filter_by(fabric_id=fabric.id).count()
            
            # Log the update
            log_audit(
                action='UPDATE',
                item_type='fabric',
                item_id=fabric.id,
                item_name=fabric.name,
                old_values=old_values,
                new_values={
                    'name': fabric.name,
                    'code': fabric.fabric_code,
                    'cost_per_yard': float(fabric.cost_per_yard) if fabric.cost_per_yard else None,
                    'vendor': fabric.fabric_vendor.name if fabric.fabric_vendor else None
                },
                affected_styles_count=affected_count
            )
            
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating fabric: {e}")
            return jsonify({'success': False, 'error': 'Failed to update fabric'}), 500
    
    elif request.method == 'DELETE':
        try:
            # Check if fabric is used in any styles
            style_fabrics = StyleFabric.query.filter_by(fabric_id=fabric_id).all()
            if style_fabrics:
                # Get the list of styles using this fabric
                styles_using = []
                for sf in style_fabrics:
                    style = Style.query.get(sf.style_id)
                    if style:
                        styles_using.append({
                            'id': style.id,
                            'vendor_style': style.vendor_style,
                            'style_name': style.style_name
                        })
                
                style_list = ', '.join([s['vendor_style'] or s['style_name'] for s in styles_using])
                return jsonify({
                    'success': False,
                    'error': f'Cannot delete: This fabric is used in {len(styles_using)} style(s): {style_list}',
                    'styles_using': styles_using
                }), 400
            
            # Capture info BEFORE delete
            fabric_name = fabric.name
            fabric_id_val = fabric.id
            old_values = {
                'name': fabric.name,
                'code': fabric.fabric_code,
                'cost_per_yard': float(fabric.cost_per_yard) if fabric.cost_per_yard else None
            }
            
            db.session.delete(fabric)
            db.session.commit()
            
            # Log the delete
            log_audit(
                action='DELETE',
                item_type='fabric',
                item_id=fabric_id_val,
                item_name=fabric_name,
                old_values=old_values
            )
            
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error deleting fabric: {e}")
            return jsonify({'success': False, 'error': 'Failed to delete fabric'}), 500

@app.route('/api/fabrics', methods=['POST'])
@role_required('admin')
def api_add_fabric():
    try:
        data = request.get_json()
        
        # Validation
        name, error = validate_required_string(data.get('name'), 'Fabric name', max_length=100)
        if error:
            return jsonify({'success': False, 'error': error}), 400
        
        cost, error = validate_positive_number(data.get('cost_per_yard'), 'Cost per yard')
        if error:
            return jsonify({'success': False, 'error': error}), 400
            
        vendor_id = data.get('fabric_vendor_id')
        if vendor_id:
            vendor_id, error = validate_positive_integer(vendor_id, 'Vendor ID')
            if error:
                return jsonify({'success': False, 'error': error}), 400
        
        # Validate fabric_code - REQUIRED and UNIQUE
        fabric_code = data.get('fabric_code', '').strip() if data.get('fabric_code') else None
        if not fabric_code:
            return jsonify({'success': False, 'error': 'Fabric code is required'}), 400
        if not fabric_code.startswith('T'):
            return jsonify({'success': False, 'error': 'Fabric code must start with T (e.g., T1, T2, T3)'}), 400
        
        # Check if fabric_code already exists
        existing_fabric = Fabric.query.filter_by(fabric_code=fabric_code).first()
        if existing_fabric:
            return jsonify({'success': False, 'error': f'Fabric code "{fabric_code}" already exists'}), 400
        
        
        fabric = Fabric(
            name=name,
            cost_per_yard=cost,
            fabric_vendor_id=vendor_id,
            fabric_code=fabric_code,
            color=data.get('color', '').strip() if data.get('color') else None
        )
        db.session.add(fabric)
        db.session.commit()
        
        # Get vendor name for logging
        vendor_name = None
        if vendor_id:
            vendor = FabricVendor.query.get(vendor_id)
            vendor_name = vendor.name if vendor else None
        
        # Log the create
        log_audit(
            action='CREATE',
            item_type='fabric',
            item_id=fabric.id,
            item_name=fabric.name,
            new_values={
                'name': fabric.name,
                'code': fabric.fabric_code,
                'cost_per_yard': float(cost),
                'vendor': vendor_name
            }
        )
        
        return jsonify({'success': True, 'id': fabric.id})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error adding fabric: {e}")
        return jsonify({'success': False, 'error': 'Failed to add fabric'}), 500


# NOTION ENDPOINTS   
@app.route('/api/notions/<int:notion_id>', methods=['GET', 'PUT', 'DELETE'])
@role_required('admin')
def api_notion_detail(notion_id):
    notion = Notion.query.get_or_404(notion_id)
    
    if request.method == 'GET':
        return jsonify({
            'id': notion.id,
            'name': notion.name,
            'cost_per_unit': notion.cost_per_unit,
            'notion_vendor_id': notion.notion_vendor_id,
            'unit_type': notion.unit_type
        })
    
    elif request.method == 'PUT':
        try:
            data = request.get_json()
            
            # Capture old values BEFORE update
            old_values = {
                'name': notion.name,
                'cost': float(notion.cost_per_unit) if notion.cost_per_unit else None,
                'vendor': notion.notion_vendor.name if notion.notion_vendor else None,
                'unit_type': notion.unit_type
            }
            
            if 'name' in data:
                name, error = validate_required_string(data.get('name'), 'Notion name', max_length=100)
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                notion.name = name
            
            if 'cost_per_unit' in data:
                cost, error = validate_positive_number(data.get('cost_per_unit'), 'Cost per unit')
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                old_cost = notion.cost_per_unit
                notion.cost_per_unit = cost
                
                # Auto-update styles if cost changed
                if old_cost != cost:
                    affected_style_ids = db.session.query(StyleNotion.style_id).filter(
                        StyleNotion.notion_id == notion.id
                    ).distinct().all()
                    
                    for (style_id,) in affected_style_ids:
                        style = Style.query.get(style_id)
                        if style and style.base_margin_percent:
                            new_cost = style.get_total_cost()
                            margin = style.base_margin_percent / 100.0
                            if margin < 1:
                                style.suggested_price = round(new_cost / (1 - margin), 2)
            
            if 'notion_vendor_id' in data:
                notion.notion_vendor_id = data.get('notion_vendor_id')
            
            if 'unit_type' in data:
                notion.unit_type = data.get('unit_type', 'each')
            
            db.session.commit()
            
            # Count affected styles
            affected_count = StyleNotion.query.filter_by(notion_id=notion.id).count()
            
            # Log the update
            log_audit(
                action='UPDATE',
                item_type='notion',
                item_id=notion.id,
                item_name=notion.name,
                old_values=old_values,
                new_values={
                    'name': notion.name,
                    'cost': float(notion.cost_per_unit) if notion.cost_per_unit else None,
                    'vendor': notion.notion_vendor.name if notion.notion_vendor else None,
                    'unit_type': notion.unit_type
                },
                affected_styles_count=affected_count
            )
            
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating notion: {e}")
            return jsonify({'success': False, 'error': 'Failed to update notion'}), 500
    
    elif request.method == 'DELETE':
        try:
            
            # Check if notion is used in any styles
            style_notions = StyleNotion.query.filter_by(notion_id=notion_id).all()
            if style_notions:
                # Get the list of styles using this notion
                styles_using = []
                for sn in style_notions:
                    style = Style.query.get(sn.style_id)
                    if style:
                        styles_using.append({
                            'id': style.id,
                            'vendor_style': style.vendor_style,
                            'style_name': style.style_name
                        })
                
                style_list = ', '.join([s['vendor_style'] or s['style_name'] for s in styles_using])
                return jsonify({
                    'success': False,
                    'error': f'Cannot delete: This notion is used in {len(styles_using)} style(s): {style_list}',
                    'styles_using': styles_using
                }), 400
            # Capture info BEFORE delete
            notion_name = notion.name
            notion_id_val = notion.id
            old_values = {
                'name': notion.name,
                'cost': float(notion.cost_per_unit) if notion.cost_per_unit else None
            }
            
            db.session.delete(notion)
            db.session.commit()
            
            # Log the delete
            log_audit(
                action='DELETE',
                item_type='notion',
                item_id=notion_id_val,
                item_name=notion_name,
                old_values=old_values
            )
            
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error deleting notion: {e}")
            return jsonify({'success': False, 'error': 'Failed to delete notion'}), 500

@app.route('/api/notions', methods=['POST'])
@role_required('admin')
def api_add_notion():
    try:
        data = request.get_json()
        
        # Validation
        name, error = validate_required_string(data.get('name'), 'Notion name', max_length=100)
        if error:
            return jsonify({'success': False, 'error': error}), 400
        
        cost, error = validate_positive_number(data.get('cost_per_unit'), 'Cost per unit')
        if error:
            return jsonify({'success': False, 'error': error}), 400
        
        vendor_id = data.get('notion_vendor_id')
        if vendor_id:
            vendor_id, error = validate_positive_integer(vendor_id, 'Vendor ID')
            if error:
                return jsonify({'success': False, 'error': error}), 400
        
        notion = Notion(
            name=name,
            cost_per_unit=cost,
            notion_vendor_id=vendor_id,
            unit_type=data.get('unit_type', 'each')
        )
        db.session.add(notion)
        db.session.commit()
        
        # Get vendor name for logging
        vendor_name = None
        if vendor_id:
            vendor = NotionVendor.query.get(vendor_id)
            vendor_name = vendor.name if vendor else None
        
        # Log the create
        log_audit(
            action='CREATE',
            item_type='notion',
            item_id=notion.id,
            item_name=notion.name,
            new_values={
                'name': notion.name,
                'cost': float(cost),
                'vendor': vendor_name,
                'unit_type': notion.unit_type
            }
        )
        
        return jsonify({'success': True, 'id': notion.id})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error adding notion: {e}")
        return jsonify({'success': False, 'error': 'Failed to add notion'}), 500

# LABOR OPERATION ENDPOINTS
@app.route('/api/labors/<int:labor_id>', methods=['GET', 'PUT', 'DELETE'])
@role_required('admin')
def api_labor_detail(labor_id):
    labor = LaborOperation.query.get_or_404(labor_id)
    
    if request.method == 'GET':
        return jsonify({
            'id': labor.id,
            'name': labor.name,
            'cost_type': labor.cost_type,
            'fixed_cost': labor.fixed_cost,
            'cost_per_hour': labor.cost_per_hour,
            'cost_per_piece': labor.cost_per_piece
        })
    
    elif request.method == 'PUT':
        try:
            data = request.get_json()
            
            if 'name' in data:
                name, error = validate_required_string(data.get('name'), 'Labor operation name', max_length=100)
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                labor.name = name
            
            if 'cost_type' in data:
                cost_type, error = validate_choice(
                    data.get('cost_type'), 
                    'Cost type', 
                    ['flat_rate', 'hourly', 'per_piece']
                )
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                labor.cost_type = cost_type
            
            if 'fixed_cost' in data:
                cost, error = validate_positive_number(data.get('fixed_cost'), 'Fixed cost', required=False)
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                old_cost = labor.fixed_cost
                labor.fixed_cost = cost
                
                # Auto-update styles if cost changed
                if old_cost != cost:
                    affected_style_ids = db.session.query(StyleLabor.style_id).filter(
                        StyleLabor.labor_operation_id == labor.id
                    ).distinct().all()
                    
                    for (style_id,) in affected_style_ids:
                        style = Style.query.get(style_id)
                        if style and style.base_margin_percent:
                            new_cost = style.get_total_cost()
                            margin = style.base_margin_percent / 100.0
                            if margin < 1:
                                style.suggested_price = round(new_cost / (1 - margin), 2)
            
            if 'cost_per_hour' in data:
                cost, error = validate_positive_number(data.get('cost_per_hour'), 'Cost per hour', required=False)
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                labor.cost_per_hour = cost
            
            if 'cost_per_piece' in data:
                cost, error = validate_positive_number(data.get('cost_per_piece'), 'Cost per piece', required=False)
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                labor.cost_per_piece = cost
            
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating labor: {e}")
            return jsonify({'success': False, 'error': 'Failed to update labor operation'}), 500
    
    elif request.method == 'DELETE':
        try:
            # Check if labor operation is used in any styles
            style_labors = StyleLabor.query.filter_by(labor_operation_id=labor_id).all()
            if style_labors:
                # Get the list of styles using this labor operation
                styles_using = []
                for sl in style_labors:
                    style = Style.query.get(sl.style_id)
                    if style:
                        styles_using.append({
                            'id': style.id,
                            'vendor_style': style.vendor_style,
                            'style_name': style.style_name
                        })
                
                style_list = ', '.join([s['vendor_style'] or s['style_name'] for s in styles_using])
                return jsonify({
                    'success': False,
                    'error': f'Cannot delete: This labor operation is used in {len(styles_using)} style(s): {style_list}',
                    'styles_using': styles_using
                }), 400
        
            
            db.session.delete(labor)
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error deleting labor: {e}")
            return jsonify({'success': False, 'error': 'Failed to delete labor operation'}), 500
@app.route('/api/labors', methods=['POST'])
@role_required('admin')
def api_add_labor():
    try:
        data = request.get_json()
        
        # Validation
        name, error = validate_required_string(data.get('name'), 'Labor operation name', max_length=100)
        if error:
            return jsonify({'success': False, 'error': error}), 400
        
        cost_type, error = validate_choice(
            data.get('cost_type'), 
            'Cost type', 
            ['flat_rate', 'hourly', 'per_piece']
        )
        if error:
            return jsonify({'success': False, 'error': error}), 400
        
        # Validate cost based on type
        fixed_cost = None
        cost_per_hour = None
        cost_per_piece = None
        
        if cost_type == 'flat_rate':
            fixed_cost, error = validate_positive_number(data.get('fixed_cost'), 'Fixed cost')
            if error:
                return jsonify({'success': False, 'error': error}), 400
        elif cost_type == 'hourly':
            cost_per_hour, error = validate_positive_number(data.get('cost_per_hour'), 'Cost per hour')
            if error:
                return jsonify({'success': False, 'error': error}), 400
        else:  # per_piece
            cost_per_piece, error = validate_positive_number(data.get('cost_per_piece'), 'Cost per piece')
            if error:
                return jsonify({'success': False, 'error': error}), 400
        
        labor = LaborOperation(
            name=name,
            cost_type=cost_type,
            fixed_cost=fixed_cost,
            cost_per_hour=cost_per_hour,
            cost_per_piece=cost_per_piece
        )
        db.session.add(labor)
        db.session.commit()
        
        return jsonify({'success': True, 'id': labor.id})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error adding labor: {e}")
        return jsonify({'success': False, 'error': 'Failed to add labor operation'}), 500


# CLEANING COST ENDPOINTS
@app.route('/api/cleanings/<int:cleaning_id>', methods=['GET', 'PUT', 'DELETE'])
@role_required('admin')
def api_cleaning_detail(cleaning_id):
    cleaning = CleaningCost.query.get_or_404(cleaning_id)
    
    if request.method == 'GET':
        return jsonify({
            'id': cleaning.id,
            'garment_type': cleaning.garment_type,
            'fixed_cost': cleaning.fixed_cost,
            'avg_minutes': cleaning.avg_minutes
        })
    
    elif request.method == 'PUT':
        try:
            data = request.get_json()
            
            if 'garment_type' in data:
                garment_type, error = validate_required_string(data.get('garment_type'), 'Garment type', max_length=50)
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                cleaning.garment_type = garment_type
            
            if 'fixed_cost' in data:
                cost, error = validate_positive_number(data.get('fixed_cost'), 'Fixed cost')
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                old_cost = cleaning.fixed_cost
                cleaning.fixed_cost = cost
                
                # Auto-update styles if cost changed
                if old_cost != cost:
                    # Find styles with this garment type
                    affected_styles = Style.query.filter_by(garment_type=cleaning.garment_type).all()
                    
                    for style in affected_styles:
                        if style.base_margin_percent:
                            new_cost = style.get_total_cost()
                            margin = style.base_margin_percent / 100.0
                            if margin < 1:
                                style.suggested_price = round(new_cost / (1 - margin), 2)
            
            if 'avg_minutes' in data:
                minutes, error = validate_positive_integer(data.get('avg_minutes'), 'Average minutes')
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                cleaning.avg_minutes = minutes
            
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating cleaning cost: {e}")
            return jsonify({'success': False, 'error': 'Failed to update cleaning cost'}), 500
    
    elif request.method == 'DELETE':
        try:
            db.session.delete(cleaning)
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error deleting cleaning cost: {e}")
            return jsonify({'success': False, 'error': 'Failed to delete cleaning cost'}), 500



@app.get("/api/cleaning-cost")
@login_required
def get_cleaning_cost():
    """Get cleaning cost for a garment type"""
    garment_type = request.args.get('type', '').strip()
    if not garment_type:
        return jsonify({"error": "type parameter required"}), 400
    
    cc = CleaningCost.query.filter_by(garment_type=garment_type).first()
    if cc:
        return jsonify({
            "garment_type": cc.garment_type,
            "fixed_cost": float(cc.fixed_cost),
            "avg_minutes": cc.avg_minutes
        })
    return jsonify({"error": "Not found"}), 404

@app.route('/api/cleanings', methods=['POST'])
@role_required('admin')
def api_add_cleaning():
    try:
        data = request.get_json()
        
        # Validation
        garment_type, error = validate_required_string(data.get('garment_type'), 'Garment type', max_length=50)
        if error:
            return jsonify({'success': False, 'error': error}), 400
        
        fixed_cost, error = validate_positive_number(data.get('fixed_cost'), 'Fixed cost')
        if error:
            return jsonify({'success': False, 'error': error}), 400
        
        avg_minutes, error = validate_positive_integer(data.get('avg_minutes'), 'Average minutes')
        if error:
            return jsonify({'success': False, 'error': error}), 400
        
        cleaning = CleaningCost(
            garment_type=garment_type,
            fixed_cost=fixed_cost,
            avg_minutes=avg_minutes
        )
        db.session.add(cleaning)
        db.session.commit()
        
        return jsonify({'success': True, 'id': cleaning.id})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error adding cleaning cost: {e}")
        return jsonify({'success': False, 'error': 'Failed to add cleaning cost'}), 500
# ===== PLACEHOLDER ROUTES FOR FUTURE FEATURES =====
@app.route("/style/new")
@login_required
@admin_required  
def style_wizard():
    # Get all master data for dropdowns
    fabric_vendors = FabricVendor.query.order_by(FabricVendor.name).all()
    notion_vendors = NotionVendor.query.order_by(NotionVendor.name).all()
    fabrics = Fabric.query.order_by(Fabric.name).all()
    notions = Notion.query.order_by(Notion.name).all()
    # Get labor operations in the order you want them displayed
    # Get labor operations in the order you want them displayed
    labor_ops=[]
    fusion = LaborOperation.query.filter(LaborOperation.name.ilike('%fus%')).first()
    if fusion: labor_ops.append(fusion)
    marker = LaborOperation.query.filter(LaborOperation.name.ilike('%marker%')).first()
    if marker: labor_ops.append(marker)
    sewing = LaborOperation.query.filter(LaborOperation.name.ilike('%sewing%')).first()
    if sewing: labor_ops.append(sewing)
    button = LaborOperation.query.filter(LaborOperation.name.ilike('%button%')).first()
    if button: labor_ops.append(button)
    garment_types = [cc.garment_type for cc in CleaningCost.query.order_by(CleaningCost.garment_type).all()]
    size_ranges = SizeRange.query.order_by(SizeRange.name).all()
    label_cost_setting = GlobalSetting.query.filter_by(setting_key='avg_label_cost').first()
    shipping_cost_setting = GlobalSetting.query.filter_by(setting_key='shipping_cost').first()
    sublimation_cost_setting = GlobalSetting.query.filter_by(setting_key='sublimation_cost').first()
    default_label_cost = label_cost_setting.setting_value if label_cost_setting else 0.20
    default_shipping_cost = shipping_cost_setting.setting_value if shipping_cost_setting else 0.00
    default_sublimation_cost = sublimation_cost_setting.setting_value if sublimation_cost_setting else 6.00

    return render_template("style_wizard.html",
                          fabric_vendors=fabric_vendors,
                          notion_vendors=notion_vendors,
                          fabrics=fabrics,
                          notions=notions,
                          labor_ops=labor_ops,
                          garment_types=garment_types,
                          size_ranges=size_ranges,
                          default_label_cost=default_label_cost,
                          default_shipping_cost=default_shipping_cost,
                          default_sublimation_cost=default_sublimation_cost)

@app.route("/style/view")
@role_required('admin', 'user')  # Both can view
def style_view():
    """View style details (read-only for users)"""
    vendor_style = request.args.get('vendor_style', '')
    style_id = request.args.get('id')
    
    # ✅ FETCH THE STYLE!
    style = None
    if vendor_style:
        style = Style.query.filter_by(vendor_style=vendor_style).first()
    elif style_id:
        style = Style.query.get(style_id)
    
    if not style:
        flash('Style not found.', 'danger')
        return redirect(url_for('view_all_styles'))
    
    # Get all master data for display
    fabric_vendors = FabricVendor.query.order_by(FabricVendor.name).all()
    notion_vendors = NotionVendor.query.order_by(NotionVendor.name).all()
    fabrics = Fabric.query.order_by(Fabric.name).all()
    notions = Notion.query.order_by(Notion.name).all()

    labor_ops = []
    fusion = LaborOperation.query.filter(LaborOperation.name.ilike('%fus%')).first()
    if fusion: labor_ops.append(fusion)
    marker = LaborOperation.query.filter(LaborOperation.name.ilike('%marker%')).first()
    if marker: labor_ops.append(marker)
    sewing = LaborOperation.query.filter(LaborOperation.name.ilike('%sewing%')).first()
    if sewing: labor_ops.append(sewing)
    button = LaborOperation.query.filter(LaborOperation.name.ilike('%button%')).first()
    if button: labor_ops.append(button)
    garment_types = [cc.garment_type for cc in CleaningCost.query.order_by(CleaningCost.garment_type).all()]
    size_ranges = SizeRange.query.order_by(SizeRange.name).all()
    
    label_cost_setting = GlobalSetting.query.filter_by(setting_key='avg_label_cost').first()
    shipping_cost_setting = GlobalSetting.query.filter_by(setting_key='shipping_cost').first()
    sublimation_cost_setting = GlobalSetting.query.filter_by(setting_key='sublimation_cost').first()
    default_label_cost = label_cost_setting.setting_value if label_cost_setting else 0.20
    default_shipping_cost = shipping_cost_setting.setting_value if shipping_cost_setting else 0.00
    default_sublimation_cost = sublimation_cost_setting.setting_value if sublimation_cost_setting else 6.00

    # Get permissions
    permissions = get_user_permissions()

    return render_template("style_wizard.html",
                          fabric_vendors=fabric_vendors,
                          notion_vendors=notion_vendors,
                          fabrics=fabrics,
                          notions=notions,
                          labor_ops=labor_ops,
                          garment_types=garment_types,
                          size_ranges=size_ranges,
                          default_label_cost=default_label_cost,
                          default_shipping_cost=default_shipping_cost,
                          default_sublimation_cost=default_sublimation_cost,
                          permissions=permissions,
                          view_mode=not permissions['can_edit'])

@app.get("/api/style/by-name")
@login_required
def api_style_by_name():
    """
    Read-only lookup by exact style_name (case-insensitive).
    Returns basics, first fabric + first notion, labor rows, and cleaning suggestion.
    """
    name = (request.args.get("name") or "").strip()
    if not name:
        return jsonify({"error": "name required"}), 400

    style = Style.query.filter(func.lower(Style.style_name) == name.lower()).first()
    if not style:
        return jsonify({"found": False}), 404

    # first fabric (if any)
    frow = (
        db.session.query(StyleFabric, Fabric)
        .join(Fabric, StyleFabric.fabric_id == Fabric.id)
        .filter(StyleFabric.style_id == style.id)
        .order_by(StyleFabric.id.asc())
        .first()
    )
    fabric_payload = None
    if frow:
        sf, f = frow
        vendor_name = f.fabric_vendor.name if f.fabric_vendor else None
        fabric_payload = {
            "vendor": vendor_name,
            "name": f.name,
            "cost_per_yard": float(f.cost_per_yard or 0),
            "yards": float(sf.yards_required or 0),
            "primary": bool(sf.is_primary or False),
        }

    # first notion (if any)
    nrow = (
        db.session.query(StyleNotion, Notion)
        .join(Notion, StyleNotion.notion_id == Notion.id)
        .filter(StyleNotion.style_id == style.id)
        .order_by(StyleNotion.id.asc())
        .first()
    )
    notion_payload = None
    if nrow:
        sn, n = nrow
        vendor_name = n.notion_vendor.name if n.notion_vendor else None
        notion_payload = {
            "vendor": vendor_name,
            "name": n.name,
            "cost_per_unit": float(n.cost_per_unit or 0),
            "qty": float(sn.quantity_required or 0),
        }

    # labor list (ordered)
    labor_rows = (
        db.session.query(StyleLabor, LaborOperation)
        .join(LaborOperation, StyleLabor.labor_operation_id == LaborOperation.id)
        .filter(StyleLabor.style_id == style.id)
        .order_by(StyleLabor.id.asc())
        .all()
    )
    labor_payload = []
    for sl, op in labor_rows:
        # Get the correct rate based on cost_type
        rate = 0
        if op.cost_type == 'flat_rate':
            rate = float(op.fixed_cost or 0)
        elif op.cost_type == 'hourly':
            rate = float(op.cost_per_hour or 0)
        elif op.cost_type == 'per_piece':
            rate = float(op.cost_per_piece or 0)
        
        # Get quantity or hours - check both time_hours and quantity
        qty_or_hours = float(sl.time_hours or 0) if sl.time_hours else float(sl.quantity or 0)
        
        labor_payload.append({
            "name": op.name,
            "cost_type": op.cost_type or "flat_rate",
            "rate": rate,
            "qty_or_hours": qty_or_hours,
        })

    # cleaning suggestion by garment_type
    cleaning_payload = None
    if style.garment_type:
        cc = CleaningCost.query.filter_by(garment_type=style.garment_type).first()
        if cc:
            cleaning_payload = {
                "type": cc.garment_type,
                "cost": float(cc.fixed_cost or 0),
                "minutes": float(cc.avg_minutes or 0),
            }

    return jsonify({
        "found": True,
        "style": {
            "vendor_style": style.vendor_style,
            "base_item_number": style.base_item_number,
            "variant_code": style.variant_code,
            "style_name": style.style_name,
            "gender": style.gender,
            "garment_type": style.garment_type,
            "size_range": style.size_range,
        },
        "fabric": fabric_payload,
        "notion": notion_payload,
        "labor": labor_payload,
        "avg_label_cost": 0.20,
        "shipping_cost": 0.00,
        "cleaning": cleaning_payload,
    }), 200

# VARIABLE ENDPOINTS
@app.route('/api/variables', methods=['GET'])
@login_required
def api_variables_get():
    """Get all variables - any authenticated user"""
    variables = Variable.query.order_by(Variable.name).all()
    return jsonify([{'id': v.id, 'name': v.name} for v in variables])


@app.route('/api/variables', methods=['POST'])
@login_required
@role_required('admin')
def api_variables_create():
    """Create new variable - admin only"""
    try:
        data = request.get_json()
        
        # Validation
        name, error = validate_required_string(data.get('name'), 'Variable name', max_length=100)
        if error:
            return jsonify({'success': False, 'error': error}), 400
        
        # Check for duplicate
        existing = Variable.query.filter(func.lower(Variable.name) == name.lower()).first()
        if existing:
            return jsonify({'success': False, 'error': 'Variable already exists'}), 400
        
        variable = Variable(name=name)
        db.session.add(variable)
        db.session.commit()
        return jsonify({'success': True, 'id': variable.id, 'name': variable.name})
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Error adding variable: {e}")
        return jsonify({'success': False, 'error': 'Failed to add variable'}), 500


@app.route('/api/variables/<int:variable_id>', methods=['GET'])
@login_required
def api_variable_get(variable_id):
    """Get variable details - any authenticated user"""
    variable = Variable.query.get_or_404(variable_id)
    return jsonify({
        'id': variable.id,
        'name': variable.name
    })


@app.route('/api/variables/<int:variable_id>', methods=['PUT', 'DELETE'])
@login_required
@role_required('admin')
def api_variable_modify(variable_id):
    """Update or delete variable - admin only"""
    variable = Variable.query.get_or_404(variable_id)
    
    if request.method == 'PUT':
        try:
            data = request.get_json()
            
            if 'name' in data:
                name, error = validate_required_string(data.get('name'), 'Variable name', max_length=100)
                if error:
                    return jsonify({'success': False, 'error': error}), 400
                variable.name = name
            
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error updating variable: {e}")
            return jsonify({'success': False, 'error': 'Failed to update variable'}), 500
    
    elif request.method == 'DELETE':
        try:
            # Check if variable is used in any styles
            style_variables = StyleVariable.query.filter_by(variable_id=variable_id).all()
            if style_variables:
                styles_using = []
                for sv in style_variables:
                    style = Style.query.get(sv.style_id)
                    if style:
                        styles_using.append({
                            'id': style.id,
                            'vendor_style': style.vendor_style,
                            'style_name': style.style_name
                        })
                
                style_list = ', '.join([s['vendor_style'] or s['style_name'] for s in styles_using])
                return jsonify({
                    'success': False,
                    'error': f'Cannot delete: This variable is used in {len(styles_using)} style(s): {style_list}',
                    'styles_using': styles_using
                }), 400
            
            db.session.delete(variable)
            db.session.commit()
            return jsonify({'success': True})
        except Exception as e:
            db.session.rollback()
            app.logger.error(f"Error deleting variable: {e}")
            return jsonify({'success': False, 'error': 'Failed to delete variable'}), 500
    

@app.route('/api/styles/search', methods=['GET'])
@limiter.limit("60 per minute")
@login_required
def search_styles():
    """Search styles by vendor_style or style_name"""
    query = request.args.get('q', '').strip()
    
    # Sanitize input
    sanitized, search_pattern = sanitize_search_query(query)
    if not sanitized:
        return jsonify([])
    
    # Search in both vendor_style and style_name
    styles = Style.query.filter(
        db.or_(
            Style.vendor_style.ilike(search_pattern, escape='\\'),
            Style.style_name.ilike(search_pattern, escape='\\')
        )
    ).order_by(Style.vendor_style).limit(20).all()
    
    return jsonify([{
        'vendor_style': s.vendor_style,
        'style_name': s.style_name
    } for s in styles])



@app.get("/api/style/by-vendor-style")
@login_required
def api_style_by_vendor_style():
    """Load style by vendor_style code"""
    vendor_style = (request.args.get("vendor_style") or "").strip()
    if not vendor_style:
        return jsonify({"error": "vendor_style required"}), 400

    style = Style.query.filter_by(vendor_style=vendor_style).first()
    if not style:
        return jsonify({"found": False}), 404

    # Get ALL fabrics - UPDATE THIS SECTION
    fabric_rows = (
        db.session.query(StyleFabric, Fabric)
        .join(Fabric, StyleFabric.fabric_id == Fabric.id)
        .filter(StyleFabric.style_id == style.id)
        .order_by(StyleFabric.id.asc())
        .all()
    )
    fabrics_payload = []
    for sf, f in fabric_rows:
        vendor_name = f.fabric_vendor.name if f.fabric_vendor else None
        fabrics_payload.append({
            "vendor": vendor_name,
            "vendor_id": f.fabric_vendor_id,
            "name": f.name,
            "fabric_id": f.id,
            "cost_per_yard": float(f.cost_per_yard or 0),
            "yards": float(sf.yards_required or 0),
            "sublimation": bool(sf.is_sublimation or False),
        })

    # Get ALL notions
    notion_rows = (
        db.session.query(StyleNotion, Notion)
        .join(Notion, StyleNotion.notion_id == Notion.id)
        .filter(StyleNotion.style_id == style.id)
        .order_by(StyleNotion.id.asc())
        .all()
    )
    notions_payload = []
    for sn, n in notion_rows:
        vendor_name = n.notion_vendor.name if n.notion_vendor else None
        notions_payload.append({
            "vendor": vendor_name,
            "vendor_id": n.notion_vendor_id,
            "name": n.name,
            "notion_id": n.id,
            "cost_per_unit": float(n.cost_per_unit or 0),
            "qty": float(sn.quantity_required or 0),
        })

    # Labor
    labor_rows = (
        db.session.query(StyleLabor, LaborOperation)
        .join(LaborOperation, StyleLabor.labor_operation_id == LaborOperation.id)
        .filter(StyleLabor.style_id == style.id)
        .order_by(StyleLabor.id.asc())
        .all()
    )
    labor_payload = []
    for sl, op in labor_rows:
        rate = 0
        if op.cost_type == 'flat_rate':
            rate = float(op.fixed_cost or 0)
        elif op.cost_type == 'hourly':
            rate = float(op.cost_per_hour or 0)
        elif op.cost_type == 'per_piece':
            rate = float(op.cost_per_piece or 0)
        
        qty_or_hours = float(sl.time_hours or 0) if sl.time_hours else float(sl.quantity or 0)
        
        labor_payload.append({
            "name": op.name,
            "rate": rate,
            "qty_or_hours": qty_or_hours,
        })

    # Cleaning
    cleaning_payload = None
    if style.garment_type:
        cc = CleaningCost.query.filter_by(garment_type=style.garment_type).first()
        if cc:
            cleaning_payload = {
                "cost": float(cc.fixed_cost or 0),
            }

    # Colors - NEW SECTION
    colors_payload = []
    color_rows = (
        db.session.query(StyleColor, Color)
        .join(Color, StyleColor.color_id == Color.id)
        .filter(StyleColor.style_id == style.id)
        .order_by(Color.name.asc())
        .all()
    )
    for sc, color in color_rows:
        colors_payload.append({
            "color_id": color.id,
            "name": color.name
        })
    
    # Variables - NEW SECTION - ADD THIS ENTIRE BLOCK
    variables_payload = []
    variable_rows = (
        db.session.query(StyleVariable, Variable)
        .join(Variable, StyleVariable.variable_id == Variable.id)
        .filter(StyleVariable.style_id == style.id)
        .order_by(Variable.name.asc())
        .all()
    )
    for sv, variable in variable_rows:
        variables_payload.append({
            "variable_id": variable.id,
            "name": variable.name
        })
    # Load from global settings instead of style record
    label_setting = GlobalSetting.query.filter_by(setting_key='avg_label_cost').first()
    shipping_setting = GlobalSetting.query.filter_by(setting_key='shipping_cost').first()

    return jsonify({
        "found": True,
        "style": {
            "id": style.id,
            "vendor_style": style.vendor_style,
            "base_item_number": style.base_item_number,
            "variant_code": style.variant_code,
            "style_name": style.style_name,
            "gender": style.gender,
            "garment_type": style.garment_type,
            "size_range": style.size_range,
            "margin": float(style.base_margin_percent or 60.0),  # NEW
            "label_cost": float(label_setting.setting_value) if label_setting else 0.20,
            "shipping_cost": float(shipping_setting.setting_value) if shipping_setting else 0.00,
            "suggested_price": float(style.suggested_price or 0) if style.suggested_price else None,  # NEW
            "notes": style.notes or '', 
        },
        "fabrics": fabrics_payload,
        "notions": notions_payload,
        "labor": labor_payload,
        "cleaning": cleaning_payload,
        "colors": colors_payload,
        "variables": variables_payload,  # NEW
}), 200


# ===== ENHANCED /api/style/save WITH FULL VALIDATION =====

@app.post("/api/style/save")
@limiter.limit("60 per minute")
@admin_required 
def api_style_save():
    """
    Save or update a Style with comprehensive validation.
    Prevents duplicates, invalid data, and ensures data integrity.
    """
    
    try:
        data = request.get_json(silent=True) or {}
        s = data.get("style") or {}
        
        # ===== STEP 1: VALIDATE REQUIRED FIELDS =====
        valid, style_name = validate_required_field(s.get("style_name"), "Style Name")
        if not valid:
            return jsonify({"error": style_name}), 400
        
        vendor_style = (s.get("vendor_style") or "").strip()
        
        # Vendor Style is REQUIRED
        if not vendor_style:
            return jsonify({"error": "Vendor Style is required"}), 400
        
        # Validate vendor_style length
        valid, vendor_style = validate_string_length(vendor_style, "Vendor Style", 50)
        if not valid:
            return jsonify({"error": vendor_style}), 400
        
        # Base Item Number is REQUIRED
        base_item_number = (s.get("base_item_number") or "").strip()
        if not base_item_number:
            return jsonify({"error": "Base Item Number is required"}), 400
        
        # Validate style_name length
        valid, style_name = validate_string_length(style_name, "Style Name", 200)
        if not valid:
            return jsonify({"error": style_name}), 400
       
        # ===== STEP 2: CHECK FOR DUPLICATES (Vendor Style only) =====
        style_id = s.get("style_id")

        # If style_id exists and is a number, this is an UPDATE
        if style_id and isinstance(style_id, (int, float)) and style_id > 0:
            # UPDATING EXISTING STYLE
            existing_style = Style.query.get(int(style_id))
            if not existing_style:
                return jsonify({"error": "Style not found"}), 404
            
            # Capture OLD values BEFORE any changes (for audit log)
            old_style_values = {
                "vendor_style": existing_style.vendor_style,
                "style_name": existing_style.style_name,
                "total_cost": str(existing_style.get_total_cost()),
                "margin": str(existing_style.base_margin_percent),
                "suggested_price": str(existing_style.suggested_price)
            }
            
            # Check if vendor_style changed and conflicts with another style
            if existing_style.vendor_style != vendor_style:
                duplicate = Style.query.filter_by(vendor_style=vendor_style).first()
                if duplicate:
                    return jsonify({"error": f"Vendor Style '{vendor_style}' already exists! Choose a different code."}), 400
            
            is_new = False
        else:
            # CREATING NEW STYLE - check vendor_style uniqueness only
            duplicate = Style.query.filter_by(vendor_style=vendor_style).first()
            if duplicate:
                return jsonify({"error": f"Vendor Style '{vendor_style}' already exists! Search and load it to edit."}), 400
            
            existing_style = Style()
            old_style_values = None  # No old values for new style
            is_new = True

        style = existing_style
                
        # ===== STEP 3: VALIDATE NUMERIC FIELDS =====
        # Validate margin
        margin = s.get("margin", 60.0)
        valid, margin = validate_percentage(margin, "Margin")
        if not valid:
            return jsonify({"error": margin}), 400
    
        # Validate suggested price - OPTIONAL (defaults to 60% margin calculation)
        suggested_price = s.get("suggested_price")
        if suggested_price:
            suggested_price, error = validate_positive_number(suggested_price, "Suggested Price", allow_zero=False)
            if error:
                return jsonify({"error": error}), 400
        else:
            suggested_price = None  # Will be calculated after costs are known
        
        # ===== STEP 4: VALIDATE FABRIC DATA =====
        fabrics_data = data.get("fabrics") or []
        for idx, f in enumerate(fabrics_data):
            if f.get("name"):
                yards = f.get("yards", 0)
                if yards:
                    yards_val, error = validate_positive_number(yards, f"Fabric #{idx+1} yards", allow_zero=False)
                    if error:
                        return jsonify({"error": error}), 400
                    f["yards"] = yards_val  # Update with validated value
        
        # ===== STEP 5: VALIDATE NOTION DATA =====
        notions_data = data.get("notions") or []
        for idx, n in enumerate(notions_data):
            if n.get("name"):
                qty = n.get("qty", 0)
                if qty:
                    qty_val, error = validate_positive_number(qty, f"Notion #{idx+1} quantity", allow_zero=False)
                    if error:
                        return jsonify({"error": error}), 400
                    n["qty"] = qty_val  # Update with validated value
                    
        # ===== STEP 5.5: VALIDATE FIRST FABRIC ROW REQUIRED =====
        valid_fabrics = [f for f in fabrics_data if f.get("name") and f.get("yards")]
        if not valid_fabrics:
            return jsonify({"error": "First fabric row is required (Vendor, Fabric, and Yards)"}), 400
        
        # ===== STEP 6: UPDATE STYLE FIELDS =====
        style.style_name = style_name
        style.vendor_style = vendor_style if vendor_style else None
        style.base_item_number = (s.get("base_item_number") or None)
        style.variant_code = (s.get("variant_code") or None)
        style.gender = (s.get("gender") or None)
        style.garment_type = (s.get("garment_type") or None)
        style.size_range = (s.get("size_range") or None)
        style.notes = (s.get("notes") or None)
        style.suggested_price = suggested_price
        db.session.add(style)
        db.session.flush()
        
        # ===== STEP 7: REPLACE EXISTING RELATIONSHIPS =====
        StyleFabric.query.filter_by(style_id=style.id).delete()
        StyleNotion.query.filter_by(style_id=style.id).delete()
        StyleLabor.query.filter_by(style_id=style.id).delete()
        StyleColor.query.filter_by(style_id=style.id).delete()
        StyleVariable.query.filter_by(style_id=style.id).delete()
        db.session.flush()
        
        # ===== STEP 8: ADD FABRICS =====
        for f in fabrics_data:
            if f.get("name"):
                fabric_name = (f["name"] or "").strip()
                vendor_name = (f.get("vendor") or "").strip()
                
                fabric = Fabric.query.filter(
                    func.lower(Fabric.name) == fabric_name.lower()
                ).first()
                
                if not fabric:
                    vendor = None
                    if vendor_name:
                        vendor = FabricVendor.query.filter(
                            func.lower(FabricVendor.name) == vendor_name.lower()
                        ).first()
                        if not vendor:
                            vendor = FabricVendor(
                                name=vendor_name, 
                                vendor_code=vendor_name[:10].upper()
                            )
                            db.session.add(vendor)
                            db.session.flush()
                    
                    fabric = Fabric(
                        name=fabric_name,
                        fabric_code=get_next_fabric_code(),
                        cost_per_yard=float(f.get("cost_per_yard") or 0),
                        fabric_vendor_id=vendor.id if vendor else None
                    )
                    db.session.add(fabric)
                    db.session.flush()
                
                sf = StyleFabric(
                    style_id=style.id,
                    fabric_id=fabric.id,
                    yards_required=float(f.get("yards") or 0),
                    is_primary=bool(f.get("primary") or False),
                    is_sublimation=bool(f.get("sublimation") or False)
                )
                db.session.add(sf)
        
        # ===== STEP 9: ADD NOTIONS =====
        for n in notions_data:
            if n.get("name"):
                notion_name = (n["name"] or "").strip()
                vendor_name = (n.get("vendor") or "").strip()
                
                notion = Notion.query.filter(
                    func.lower(Notion.name) == notion_name.lower()
                ).first()
                
                if not notion:
                    vendor = None
                    if vendor_name:
                        vendor = NotionVendor.query.filter(
                            func.lower(NotionVendor.name) == vendor_name.lower()
                        ).first()
                        if not vendor:
                            vendor = NotionVendor(
                                name=vendor_name, 
                                vendor_code=vendor_name[:10].upper()
                            )
                            db.session.add(vendor)
                            db.session.flush()
                    
                    notion = Notion(
                        name=notion_name,
                        cost_per_unit=float(n.get("cost_per_unit") or 0),
                        unit_type='each',
                        notion_vendor_id=vendor.id if vendor else None
                    )
                    db.session.add(notion)
                    db.session.flush()
                
                sn = StyleNotion(
                    style_id=style.id,
                    notion_id=notion.id,
                    quantity_required=float(n.get("qty") or 0)
                )
                db.session.add(sn)
        
        # ===== STEP 10: ADD LABOR =====
        for l in data.get("labor") or []:
            if not l.get("name"):
                continue
            
            op = LaborOperation.query.filter(
                func.lower(LaborOperation.name) == l["name"].strip().lower()
            ).first()
            
            if not op:
                continue
            
            qty_or_hours = float(l.get("qty_or_hours") or 0)
            
            # Validate labor hours/quantity
            if qty_or_hours < 0:
                return jsonify({"error": f"Labor {l['name']} cannot have negative hours/quantity"}), 400
            
            if op.cost_type == 'hourly':
                sl = StyleLabor(
                    style_id=style.id,
                    labor_operation_id=op.id,
                    time_hours=qty_or_hours,
                    quantity=0
                )
            else:
                sl = StyleLabor(
                    style_id=style.id,
                    labor_operation_id=op.id,
                    time_hours=0,
                    quantity=int(qty_or_hours) if qty_or_hours else 0
                )
            db.session.add(sl)
        
        # ===== STEP 11: ADD COLORS =====
        for c in data.get("colors") or []:
            if c.get("name"):
                color_name = (c["name"] or "").strip()
                color = Color.query.filter(
                    func.lower(Color.name) == color_name.lower()
                ).first()
                
                if not color:
                    color = Color(name=color_name)
                    db.session.add(color)
                    db.session.flush()
                
                sc = StyleColor(style_id=style.id, color_id=color.id)
                db.session.add(sc)
        
        # ===== STEP 12: ADD VARIABLES =====
        for v in data.get("variables") or []:
            if v.get("name"):
                var_name = (v["name"] or "").strip()
                variable = Variable.query.filter(
                    func.lower(Variable.name) == var_name.lower()
                ).first()
                
                if not variable:
                    variable = Variable(name=var_name)
                    db.session.add(variable)
                    db.session.flush()
                
                sv = StyleVariable(style_id=style.id, variable_id=variable.id)
                db.session.add(sv)
        
        # ===== STEP 12.5: RECALCULATE MARGIN WITH ACTUAL COSTS =====
        db.session.flush()  # Ensure all relationships are saved
        total_cost = style.get_total_cost()
        
        if total_cost > 0:
            # Use the margin from payload (user-set or default 60%)
            style.base_margin_percent = margin if margin else 60.0
            # Calculate sale price from margin
            margin_decimal = style.base_margin_percent / 100.0
            if margin_decimal >= 0.99:  # Prevent division by zero or near-zero
                margin_decimal = 0.99     
            style.suggested_price = round(total_cost / (1 - margin_decimal), 2)
        else:
            # No costs yet - use defaults
            style.base_margin_percent = margin if margin else 60.0
            style.suggested_price = suggested_price if suggested_price else 0

        # ===== STEP 13: COMMIT ALL CHANGES =====

        db.session.commit()

        # ===== STEP 14: LOG AUDIT =====
        try:
            if is_new:
                # Log new style creation
                new_values = {
                    "vendor_style": style.vendor_style,
                    "style_name": style.style_name,
                    "total_cost": str(style.get_total_cost()),
                    "fabrics": [f.get("name") for f in fabrics_data if f.get("name")],
                    "notions": [n.get("name") for n in notions_data if n.get("name")]
                }
                log_audit(
                    action="CREATE",
                    item_type="style",
                    item_id=style.id,
                    item_name=style.vendor_style,
                    old_values=None,
                    new_values=new_values,
                    details=f"Created style: {style.style_name}"
                )
            else:
                # Log style update with changes
                new_values = {
                    "vendor_style": style.vendor_style,
                    "style_name": style.style_name,
                    "total_cost": str(style.get_total_cost()),
                    "margin": str(style.base_margin_percent),
                    "suggested_price": str(style.suggested_price)
                }
                log_audit(
                    action="UPDATE",
                    item_type="style",
                    item_id=style.id,
                    item_name=style.vendor_style,
                    old_values=old_style_values,
                    new_values=new_values,
                    details=f"Updated style: {style.style_name}"
                )
        except Exception as e:
            app.logger.error(f"Failed to log audit for style: {e}")

        # Return appropriate message
        if is_new:
            message = "✅ New style created successfully!"
        else:
            message = "✅ Style updated successfully!"

        return jsonify({
            "success": True,
            "new": is_new,
            "style_id": style.id,
            "message": message
        }), 200
    
    except IntegrityError as e:
        db.session.rollback()
        app.logger.warning(f"IntegrityError in style save: {e}")
        # Parse the error to give user-friendly message
        error_msg = str(e.orig).lower() if e.orig else str(e).lower()
        if 'vendor_style' in error_msg:
            return jsonify({"success": False, "error": "A style with this Vendor Style already exists. Please use a different code."}), 400
        else:
            return jsonify({"success": False, "error": "This record already exists. Please check for duplicates."}), 400
    
    except ValueError as e:
        db.session.rollback()
        return jsonify({"success": False, "error": f"Invalid data: {str(e)}"}), 400
    
    except Exception as e:
        db.session.rollback()
        app.logger.error(f"Style save error: {e}")
        return jsonify({"success": False, "error": f"Save failed: {str(e)}"}), 500

# ===== END OF ENHANCED api_style_save =====
    
@app.get("/api/style/search")
@limiter.limit("60 per minute")
@login_required
def api_style_search():
    """Search styles by name - with sanitized input"""
    q = request.args.get("q", "").strip()
    
    # Sanitize input
    sanitized, search_pattern = sanitize_search_query(q)
    if not sanitized:
        return jsonify([])
    
    rows = (Style.query
            .filter(Style.style_name.ilike(search_pattern, escape='\\'))
            .order_by(Style.style_name.asc())
            .limit(20).all())
    return jsonify([r.style_name for r in rows])



@app.route('/download-import-template')
@admin_required
def download_import_template():
    """Download Excel template for importing styles"""
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "SAMPLE-001 Style Name"
    
    # Define styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    section_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
    section_font = Font(bold=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Row 1: Title
    ws['A1'] = "J.A. Uniforms Standard Cost Analysis"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')
    
    # Row 2: Headers
    headers = ['Item#', 'Name', 'Variant', 'Fabric', 'Minimum Qty']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
    
    # Row 3: Sample Style Data
    sample_data = ['44217', 'Ladies Ls "Server" Jacket', 'Base', 'T6', 24]
    for col, value in enumerate(sample_data, 1):
        cell = ws.cell(row=3, column=col, value=value)
        cell.border = thin_border
    
    # Row 4: MATERIALS Header
    ws['A4'] = "MATERIALS"
    ws['B4'] = "DESCRIPTION"
    ws['C4'] = "COST"
    ws['D4'] = "QTY"
    ws['E4'] = "TOTAL"
    for col in range(1, 6):
        cell = ws.cell(row=4, column=col)
        cell.fill = section_fill
        cell.font = section_font
        cell.border = thin_border
    
    # Row 5: Fabric
    ws['A5'] = "Fabric"
    ws['B5'] = "T4 Perfectly Poplin"
    ws['C5'] = 4
    ws['D5'] = 1.89
    ws['E5'] = "=C5*D5"
    
    # Row 6: Fabric#2
    ws['A6'] = "Fabric#2"
    
    # Row 7: Lining
    ws['A7'] = "Lining"
    
    # Row 9: Buttons/Notions
    ws['A9'] = "Buttons"
    ws['B9'] = "30L"
    ws['C9'] = 0.05
    ws['D9'] = 5
    ws['E9'] = "=C9*D9"
    
    # Row 11: Shoulder Pads
    ws['A11'] = "Shoulder Pads"
    ws['B11'] = "BG Lieberman"
    ws['C11'] = 1.04
    ws['D11'] = 2
    ws['E11'] = "=C11*D11"
    
    # Row 12: Labels
    ws['A12'] = "Labels"
    ws['B12'] = "Average Cost for Size/Care/JAU"
    ws['C12'] = 0.2
    ws['D12'] = 1
    ws['E12'] = "=C12*D12"
    
    # Row 13: Total Materials
    ws['C13'] = "Total Materials Cost"
    ws['E13'] = "=SUM(E5:E12)"
    ws['C13'].font = Font(bold=True)
    
    # Row 14: LABOR Header
    ws['A14'] = "LABOR"
    ws['B14'] = "DESCRIPTION"
    ws['C14'] = "COST"
    ws['D14'] = "QTY"
    ws['E14'] = "TOTAL"
    for col in range(1, 6):
        cell = ws.cell(row=14, column=col)
        cell.fill = section_fill
        cell.font = section_font
        cell.border = thin_border
    
    # Row 15-17: Marker/Cut options
    ws['A15'] = "Simple Marker/Cut"
    ws['C15'] = 1.5
    
    ws['A16'] = "Marker/Cut/Fusing"
    ws['B16'] = "Average M/C/F Cost"
    ws['C16'] = 3
    ws['D16'] = 1
    ws['E16'] = "=C16*D16"
    
    ws['A17'] = "Additional Cut"
    ws['C17'] = 1.5
    
    # Row 19: Sewing
    ws['A19'] = "Sewing"
    ws['B19'] = "Seamstress Average Hourly Cost"
    ws['C19'] = 19.32
    ws['D19'] = 0.96
    ws['E19'] = "=C19*D19"
    
    # Row 21-23: Finishing options
    ws['A21'] = "Finishing"
    ws['B21'] = "Apron/Simple"
    ws['C21'] = 0.75
    
    ws['A22'] = "Finishing"
    ws['B22'] = "Shirts/Pants"
    ws['C22'] = 1.25
    
    ws['A23'] = "Finishing"
    ws['B23'] = "Jackets/Dress"
    ws['C23'] = 2
    ws['D23'] = 1
    ws['E23'] = "=C23*D23"
    
    # Row 24: Total Labor
    ws['C24'] = "Total Labor Cost"
    ws['E24'] = "=SUM(E15:E23)"
    ws['C24'].font = Font(bold=True)
    
    # Row 26: Total Cost
    ws['A26'] = "IMAGE"
    ws['C26'] = "TOTAL COST"
    ws['E26'] = "=E13+E24"
    ws['C26'].font = Font(bold=True)
    ws['E26'].font = Font(bold=True, size=12)
    
    # Row 28-29: Margins
    ws['C28'] = "Price @ 50% Margin"
    ws['E28'] = "=E26/0.5"
    ws['C29'] = "Price @ 60% Margin"
    ws['E29'] = "=E26/0.4"
    
    # Row 30-31: Sales Price
    ws['D30'] = "Sales Price"
    ws['E30'] = "Margin"
    ws['D31'] = 115.95
    ws['E31'] = "=(D31-E26)/D31"
    
    # Row 32-33: Date
    ws['C32'] = "Date:"
    ws['C33'] = "2024-07-02"
    
    # Row 35-36: Size Range
    ws['C35'] = "Size Range:"
    ws['C36'] = "XXS-4XL (Alpha)"
    
    # Row 38-39: Clients
    ws['C38'] = "Clients:"
    ws['C39'] = "The Setai"
    
    # Row 43-44: ALVA/Photographed
    ws['C43'] = "ALVA"
    ws['D43'] = "Photographed"
    ws['E43'] = "Old#"
    ws['C44'] = "YES"
    ws['D44'] = "NO"
    
    # Row 45: Notes
    ws['A45'] = "NOTES"
    
    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 15
    
    wb.save(output)
    output.seek(0)
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name='ja_uniforms_import_template.xlsx'
    )


@app.route('/import-excel', methods=['GET', 'POST'])
@admin_required
def import_excel():
    """Import styles from Excel file - One sheet per style format"""
    
    if request.method == 'GET':
        return render_template('import_excel.html')
    
    # POST - Process the import
    if 'excel_file' not in request.files:
        flash('No file selected', 'error')
        return redirect(url_for('import_excel'))
    
    file = request.files['excel_file']
    if file.filename == '':
        flash('No file selected', 'error')
        return redirect(url_for('import_excel'))
    
    # Get uploaded images
    uploaded_images = request.files.getlist('images')
    image_dict = {}
    for img in uploaded_images:
        if img.filename:
            # Store by filename (lowercase for matching)
            image_dict[img.filename.lower()] = img
    
    try:
        # Read Excel file with all sheets
        xl = pd.ExcelFile(file)
        sheet_names = xl.sheet_names
        
        imported_count = 0
        updated_count = 0
        skipped_count = 0
        errors = []
        
        # Ensure upload folder exists
        upload_folder = os.path.join(app.root_path, 'static', 'uploads', 'styles')
        os.makedirs(upload_folder, exist_ok=True)
        
        # Process each sheet (each sheet = one style)
        for sheet_name in sheet_names:
            try:
                # Read sheet without headers
                df = pd.read_excel(xl, sheet_name=sheet_name, header=None)
                
                # Extract style info from Row 3 (index 2)
                # Format: [Item#, Name, Variant, Fabric Code, Min Qty]
                if len(df) < 3:
                    errors.append(f"Sheet '{sheet_name}': Not enough rows")
                    continue
                
                item_number = str(df.iloc[2, 0]).strip() if pd.notna(df.iloc[2, 0]) else ""
                style_name = str(df.iloc[2, 1]).strip() if pd.notna(df.iloc[2, 1]) else ""
                variant = str(df.iloc[2, 2]).strip() if pd.notna(df.iloc[2, 2]) else "Base"
                fabric_code = str(df.iloc[2, 3]).strip() if pd.notna(df.iloc[2, 3]) else ""
                
                # Skip if no item number
                if not item_number or item_number == 'nan':
                    errors.append(f"Sheet '{sheet_name}': Missing Item#")
                    continue
                
                # Create vendor_style
                if variant and variant != 'nan' and variant != 'Base':
                    vendor_style = f"{item_number}-{variant}"
                else:
                    vendor_style = item_number
                
                # Determine gender from style name
                gender = "UNISEX"
                style_name_upper = style_name.upper()
                if "LADIES" in style_name_upper or "WOMEN" in style_name_upper:
                    gender = "LADIES"
                elif "MENS" in style_name_upper or "MEN'S" in style_name_upper:
                    gender = "MENS"
                
                # Determine garment type from style name
                garment_type = "SS TOP/ SS DRESS"  # Default
                if "JACKET" in style_name_upper:
                    garment_type = "SS JACKET/LINED SS DRESS" if "LS" not in style_name_upper else "LS JACKET/LINED LS DRESS"
                elif "DRESS" in style_name_upper:
                    garment_type = "SS TOP/ SS DRESS" if "LS" not in style_name_upper else "LS TOP/ LS DRESS"
                elif "PANTS" in style_name_upper or "TROUSER" in style_name_upper:
                    garment_type = "PANTS"
                elif "SHIRT" in style_name_upper or "TOP" in style_name_upper or "BLOUSE" in style_name_upper:
                    garment_type = "SS TOP/ SS DRESS" if "LS" not in style_name_upper else "LS TOP/ LS DRESS"
                elif "VEST" in style_name_upper:
                    garment_type = "VEST"
                elif "APRON" in style_name_upper:
                    garment_type = "APRON"
                elif "SKIRT" in style_name_upper or "SHORT" in style_name_upper:
                    garment_type = "SHORTS/SKIRTS"
                
                # Extract size range from Row 36 (index 35)
                size_range = "XS-4XL"  # Default
                if len(df) > 35 and pd.notna(df.iloc[35, 2]):
                    size_range = str(df.iloc[35, 2]).strip()
                    if size_range == 'nan':
                        size_range = "XS-4XL"
                
                # Check if style already exists
                existing_style = Style.query.filter_by(vendor_style=vendor_style).first()
                
                if existing_style:
                    # Update existing style
                    existing_style.style_name = style_name
                    existing_style.gender = gender
                    existing_style.garment_type = garment_type
                    existing_style.size_range = size_range
                    existing_style.updated_at = datetime.now()
                    current_style = existing_style
                    updated_count += 1
                else:
                    # Create new style
                    current_style = Style(
                        vendor_style=vendor_style,
                        base_item_number=item_number,
                        variant_code=variant if variant and variant != 'nan' and variant != 'Base' else None,
                        style_name=style_name,
                        gender=gender,
                        garment_type=garment_type,
                        size_range=size_range
                    )
                    db.session.add(current_style)
                    db.session.flush()
                    imported_count += 1
                
                # ===== PROCESS MATERIALS =====
                
                # Fabric (Row 5, index 4)
                if len(df) > 4 and pd.notna(df.iloc[4, 0]) and str(df.iloc[4, 0]).strip().upper() == 'FABRIC':
                    fabric_name = str(df.iloc[4, 1]).strip() if pd.notna(df.iloc[4, 1]) else None
                    fabric_cost = float(df.iloc[4, 2]) if pd.notna(df.iloc[4, 2]) else 6.00
                    fabric_yards = float(df.iloc[4, 3]) if pd.notna(df.iloc[4, 3]) else 1.5
                    
                    if fabric_name and fabric_name != 'nan':
                        # Find or create fabric
                        fabric = Fabric.query.filter_by(name=fabric_name).first()
                        if not fabric:
                            # Get or create IMPORTED vendor
                            vendor = FabricVendor.query.filter_by(name="IMPORTED").first()
                            if not vendor:
                                vendor = FabricVendor(name="IMPORTED", vendor_code="IMP")
                                db.session.add(vendor)
                                db.session.flush()
                            
                            fabric = Fabric(
                                name=fabric_name,
                                fabric_code=get_next_fabric_code(),
                                cost_per_yard=fabric_cost,
                                fabric_vendor_id=vendor.id
                            )
                            db.session.add(fabric)
                            db.session.flush()
                        
                        # Check if relationship exists
                        existing_sf = StyleFabric.query.filter_by(
                            style_id=current_style.id,
                            fabric_id=fabric.id
                        ).first()
                        
                        if existing_sf:
                            # Update existing
                            existing_sf.yards_required = fabric_yards
                            existing_sf.is_primary = True
                        else:
                            style_fabric = StyleFabric(
                                style_id=current_style.id,
                                fabric_id=fabric.id,
                                yards_required=fabric_yards,
                                is_primary=True
                            )
                            db.session.add(style_fabric)

                    
                
                # Fabric#2 (Row 6, index 5)
                if len(df) > 5 and pd.notna(df.iloc[5, 1]):
                    fabric2_name = str(df.iloc[5, 1]).strip()
                    fabric2_cost = float(df.iloc[5, 2]) if pd.notna(df.iloc[5, 2]) else 6.00
                    fabric2_yards = float(df.iloc[5, 3]) if pd.notna(df.iloc[5, 3]) else 1.0
                    
                    if fabric2_name and fabric2_name != 'nan':
                        fabric2 = Fabric.query.filter_by(name=fabric2_name).first()
                        if not fabric2:
                            vendor = FabricVendor.query.filter_by(name="IMPORTED").first()
                            fabric2 = Fabric(
                                name=fabric2_name,
                                fabric_code=get_next_fabric_code(),
                                cost_per_yard=fabric2_cost,
                                fabric_vendor_id=vendor.id
                            )
                            db.session.add(fabric2)
                            db.session.flush()
                        
                        existing_sf2 = StyleFabric.query.filter_by(
                            style_id=current_style.id,
                            fabric_id=fabric2.id
                        ).first()
                        
                        if existing_sf2:
                            existing_sf2.yards_required = fabric2_yards
                            existing_sf2.is_primary = False
                        else:
                            style_fabric2 = StyleFabric(
                                style_id=current_style.id,
                                fabric_id=fabric2.id,
                                yards_required=fabric2_yards,
                                is_primary=False
                            )
                            db.session.add(style_fabric2)
                
                # Lining (Row 7, index 6)
                if len(df) > 6 and pd.notna(df.iloc[6, 1]):
                    lining_name = str(df.iloc[6, 1]).strip()
                    lining_cost = float(df.iloc[6, 2]) if pd.notna(df.iloc[6, 2]) else 4.00
                    lining_yards = float(df.iloc[6, 3]) if pd.notna(df.iloc[6, 3]) else 1.0
                    
                    if lining_name and lining_name != 'nan':
                        lining = Fabric.query.filter_by(name=lining_name).first()
                        if not lining:
                            vendor = FabricVendor.query.filter_by(name="IMPORTED").first()
                            lining = Fabric(
                                name=lining_name,
                                fabric_code=get_next_fabric_code(),
                                cost_per_yard=lining_cost,
                                fabric_vendor_id=vendor.id
                            )
                            db.session.add(lining)
                            db.session.flush()
                        
                        existing_lining = StyleFabric.query.filter_by(
                            style_id=current_style.id,
                            fabric_id=lining.id
                        ).first()
                        
                        if not existing_lining:
                            style_lining = StyleFabric(
                                style_id=current_style.id,
                                fabric_id=lining.id,
                                yards_required=lining_yards,
                                is_primary=False
                            )
                            db.session.add(style_lining)
                
                # Buttons (Row 9, index 8)
                if len(df) > 8 and pd.notna(df.iloc[8, 0]):
                    notion_type = str(df.iloc[8, 0]).strip()
                    if notion_type and notion_type != 'nan':
                        notion_name = str(df.iloc[8, 1]).strip() if pd.notna(df.iloc[8, 1]) else notion_type
                        notion_cost = float(df.iloc[8, 2]) if pd.notna(df.iloc[8, 2]) else 0.05
                        notion_qty = float(df.iloc[8, 3]) if pd.notna(df.iloc[8, 3]) else 1
                        
                        if notion_name and notion_name != 'nan' and notion_qty > 0:
                            notion = Notion.query.filter_by(name=notion_name).first()
                            if not notion:
                                n_vendor = NotionVendor.query.filter_by(name="IMPORTED").first()
                                if not n_vendor:
                                    n_vendor = NotionVendor(name="IMPORTED", vendor_code="IMP")
                                    db.session.add(n_vendor)
                                    db.session.flush()
                                
                                notion = Notion(
                                    name=notion_name,
                                    cost_per_unit=notion_cost,
                                    notion_vendor_id=n_vendor.id
                                )
                                db.session.add(notion)
                                db.session.flush()
                            
                            existing_sn = StyleNotion.query.filter_by(
                                style_id=current_style.id,
                                notion_id=notion.id
                            ).first()
                            
                            if existing_sn:
                                existing_sn.quantity_required = notion_qty
                            else:
                                style_notion = StyleNotion(
                                    style_id=current_style.id,
                                    notion_id=notion.id,
                                    quantity_required=notion_qty
                                )
                                db.session.add(style_notion)
                    
                # Shoulder Pads (Row 11, index 10)
                if len(df) > 10 and pd.notna(df.iloc[10, 0]):
                    sp_type = str(df.iloc[10, 0]).strip()
                    if sp_type and sp_type != 'nan' and 'SHOULDER' in sp_type.upper():
                        sp_name = str(df.iloc[10, 1]).strip() if pd.notna(df.iloc[10, 1]) else "Shoulder Pads"
                        sp_cost = float(df.iloc[10, 2]) if pd.notna(df.iloc[10, 2]) else 1.00
                        sp_qty = int(float(df.iloc[10, 3])) if pd.notna(df.iloc[10, 3]) else 2
                        
                        if sp_name and sp_name != 'nan' and sp_qty > 0:
                            sp_notion = Notion.query.filter_by(name=sp_name).first()
                            if not sp_notion:
                                n_vendor = NotionVendor.query.filter_by(name="IMPORTED").first()
                                sp_notion = Notion(
                                    name=sp_name,
                                    cost_per_unit=sp_cost,
                                    notion_vendor_id=n_vendor.id
                                )
                                db.session.add(sp_notion)
                                db.session.flush()
                            
                            existing_sp = StyleNotion.query.filter_by(
                                style_id=current_style.id,
                                notion_id=sp_notion.id
                            ).first()
                            
                            if not existing_sp:
                                style_sp = StyleNotion(
                                    style_id=current_style.id,
                                    notion_id=sp_notion.id,
                                    quantity_required=sp_qty
                                )
                                db.session.add(style_sp)
                
                # ===== PROCESS LABOR =====
                
                # Marker/Cut/Fusing (Row 16, index 15)
                if len(df) > 15:
                    mcf_label = str(df.iloc[15, 0]).strip() if pd.notna(df.iloc[15, 0]) else ""
                    if 'MARKER' in mcf_label.upper() or 'FUSING' in mcf_label.upper():
                        mcf_cost = float(df.iloc[15, 2]) if pd.notna(df.iloc[15, 2]) else 3.00
                        mcf_qty = float(df.iloc[15, 3]) if pd.notna(df.iloc[15, 3]) else 1
                        
                        if mcf_qty > 0:
                            # Find Marker+Cut labor operation
                            labor_op = LaborOperation.query.filter(
                                LaborOperation.name.ilike('%marker%cut%')
                            ).first()
                            
                            if labor_op:
                                existing_sl = StyleLabor.query.filter_by(
                                    style_id=current_style.id,
                                    labor_operation_id=labor_op.id
                                ).first()
                                
                                if not existing_sl:
                                    style_labor = StyleLabor(
                                        style_id=current_style.id,
                                        labor_operation_id=labor_op.id,
                                        quantity=int(mcf_qty)
                                    )
                                    db.session.add(style_labor)
                
                # Sewing (Row 19, index 18)
                if len(df) > 18:
                    sewing_label = str(df.iloc[18, 0]).strip() if pd.notna(df.iloc[18, 0]) else ""
                    if 'SEWING' in sewing_label.upper():
                        sewing_hours = float(df.iloc[18, 3]) if pd.notna(df.iloc[18, 3]) else 0
                        
                        if sewing_hours > 0:
                            # Find Sewing labor operation
                            sewing_op = LaborOperation.query.filter(
                                LaborOperation.name.ilike('%sewing%')
                            ).first()
                            
                            if sewing_op:
                                existing_sewing = StyleLabor.query.filter_by(
                                    style_id=current_style.id,
                                    labor_operation_id=sewing_op.id
                                ).first()
                                
                                if not existing_sewing:
                                    style_sewing = StyleLabor(
                                        style_id=current_style.id,
                                        labor_operation_id=sewing_op.id,
                                        time_hours=sewing_hours
                                    )
                                    db.session.add(style_sewing)
                
                # Finishing - Check rows 21-23 (index 20-22)
                for finish_row in [20, 21, 22]:
                    if len(df) > finish_row:
                        finish_label = str(df.iloc[finish_row, 0]).strip() if pd.notna(df.iloc[finish_row, 0]) else ""
                        if 'FINISH' in finish_label.upper():
                            finish_qty = float(df.iloc[finish_row, 3]) if pd.notna(df.iloc[finish_row, 3]) else 0
                            
                            if finish_qty > 0:
                                # Use the garment type for cleaning/ironing
                                cleaning_op = CleaningCost.query.filter_by(garment_type=garment_type).first()
                                if cleaning_op:
                                    # Add cleaning cost as labor
                                    pass  # Cleaning costs are usually automatic based on garment type
                                break  # Only process one finishing entry
                
                # Commit after each style
                db.session.commit()
                
            except Exception as e:
                db.session.rollback()
                errors.append(f"Sheet '{sheet_name}': {str(e)}")
                continue
        
        # Flash success message
        if imported_count > 0 or updated_count > 0:
            flash(f'Successfully imported {imported_count} new styles and updated {updated_count} existing styles!', 'success')
        
        if errors:
            flash(f'{len(errors)} sheets had issues. Check the details below.', 'warning')
        
        # Return results page
        return render_template('import_results.html',
                             imported_count=imported_count,
                             updated_count=updated_count,
                             skipped_count=skipped_count,
                             errors=errors)
    
    except Exception as e:
        db.session.rollback()
        flash(f'Import failed: {str(e)}', 'error')
        return redirect(url_for('import_excel'))




# ===== APPLICATION STARTUP =====
if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    
    # Get debug mode from environment variable (defaults to False)
    debug_mode = os.environ.get('FLASK_DEBUG', 'False').lower() == 'true'
    
    # Safety check: Never allow debug in production
    if debug_mode and os.environ.get('FLASK_ENV') == 'production':
        raise ValueError("❌ CRITICAL: Cannot run with debug=True in production!")
    
    # Show warning if debug is enabled
    if debug_mode:
        app.logger.info("\n" + "="*70)
        app.logger.info("⚠️  WARNING: Running in DEBUG mode!")
        app.logger.info("⚠️  This is for development only - NEVER use in production!")
        app.logger.info("⚠️  Set FLASK_DEBUG=False in .env for production")
        app.logger.info("="*70 + "\n")
    
    app.run(debug=debug_mode, host='0.0.0.0', port=5000)